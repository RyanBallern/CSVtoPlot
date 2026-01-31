"""Export statistical results as formatted tables."""

import pandas as pd
from pathlib import Path
from typing import Dict, List
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from ..processors.statistics import StatisticsEngine


class StatisticsTableExporter:
    """Exports comprehensive statistical results as formatted tables."""

    def __init__(self, stats_engine: StatisticsEngine):
        """Initialize statistics table exporter.

        Args:
            stats_engine: Statistics engine for calculations
        """
        self.stats = stats_engine

    def create_statistics_tables(self, data: Dict[str, pd.Series],
                                  parameter_name: str) -> Dict[str, pd.DataFrame]:
        """Create comprehensive statistical tables.

        Args:
            data: Dictionary mapping condition names to data series
            parameter_name: Name of the parameter

        Returns:
            Dict with 'summary', 'anova', 'pairwise' DataFrames
        """
        tables = {}

        # 1. Summary statistics table
        tables['summary'] = self._create_summary_table(data, parameter_name)

        # 2. ANOVA results table
        tables['anova'] = self._create_anova_table(data, parameter_name)

        # 3. Pairwise comparisons table
        tables['pairwise'] = self._create_pairwise_table(data, parameter_name)

        return tables

    def _create_summary_table(self, data: Dict[str, pd.Series],
                              parameter_name: str) -> pd.DataFrame:
        """Create summary statistics table with values.

        Args:
            data: Dictionary mapping condition names to data series
            parameter_name: Name of the parameter

        Returns:
            DataFrame with summary statistics
        """
        import numpy as np
        rows = []

        for condition, series in data.items():
            # Calculate summary statistics
            n = len(series)
            mean = series.mean()
            std = series.std(ddof=1)
            sem = std / np.sqrt(n) if n > 0 else 0
            median = series.median()
            min_val = series.min()
            max_val = series.max()
            q25 = series.quantile(0.25)
            q75 = series.quantile(0.75)

            rows.append({
                'Parameter': parameter_name,
                'Condition': condition,
                'N': n,
                'Mean': f"{mean:.3f}",
                'SEM': f"{sem:.3f}",
                'SD': f"{std:.3f}",
                'Median': f"{median:.3f}",
                'Min': f"{min_val:.3f}",
                'Max': f"{max_val:.3f}",
                'Q25': f"{q25:.3f}",
                'Q75': f"{q75:.3f}"
            })

        return pd.DataFrame(rows)

    def _create_anova_table(self, data: Dict[str, pd.Series],
                            parameter_name: str) -> pd.DataFrame:
        """Create ANOVA results table.

        Args:
            data: Dictionary mapping condition names to data series
            parameter_name: Name of the parameter

        Returns:
            DataFrame with ANOVA results
        """
        # Convert data to DataFrame format for auto_compare
        df_data = []
        for condition, series in data.items():
            for value in series:
                df_data.append({'condition': condition, 'value': value})
        df = pd.DataFrame(df_data)

        # Run auto_compare to get statistical results
        result = self.stats.auto_compare(df, 'value', 'condition')

        rows = [{
            'Parameter': parameter_name,
            'Test': result['main_test'].test_name,
            'F-statistic': f"{result['main_test'].statistic:.4f}",
            'p-value': f"{result['main_test'].p_value:.4e}",
            'Significant': 'Yes' if result['main_test'].significant else 'No',
            'Alpha': self.stats.alpha
        }]

        # Add normality test results if available
        if result.get('normality_tests'):
            for condition, norm_result in result['normality_tests'].items():
                rows.append({
                    'Parameter': parameter_name,
                    'Test': f'Normality ({condition})',
                    'F-statistic': '-',
                    'p-value': f"{norm_result.p_value:.4e}",
                    'Significant': 'Normal' if norm_result.is_normal else 'Non-normal',
                    'Alpha': self.stats.alpha
                })

        return pd.DataFrame(rows)

    def _create_pairwise_table(self, data: Dict[str, pd.Series],
                               parameter_name: str) -> pd.DataFrame:
        """Create pairwise comparisons table.

        Args:
            data: Dictionary mapping condition names to data series
            parameter_name: Name of the parameter

        Returns:
            DataFrame with pairwise comparison results
        """
        # Convert data to DataFrame format for auto_compare
        df_data = []
        for condition, series in data.items():
            for value in series:
                df_data.append({'condition': condition, 'value': value})
        df = pd.DataFrame(df_data)

        # Run auto_compare to get statistical results
        result = self.stats.auto_compare(df, 'value', 'condition')

        # Get post-hoc tests if available
        post_hoc_tests = result.get('post_hoc_tests')

        if not post_hoc_tests:
            return pd.DataFrame({
                'Parameter': [parameter_name],
                'Group 1': ['No significant'],
                'Group 2': ['differences found'],
                'Mean Difference': ['-'],
                'p-value': ['-'],
                'Significant': ['No']
            })

        rows = []
        for test in post_hoc_tests:
            rows.append({
                'Parameter': parameter_name,
                'Group 1': test.group1,
                'Group 2': test.group2,
                'Mean Difference': f"{test.mean_diff:.3f}",
                'p-value': f"{test.p_value:.4e}",
                'Significant': self._get_significance_level(test.p_value)
            })

        return pd.DataFrame(rows)

    def _get_significance_level(self, p_value: float) -> str:
        """Convert p-value to significance level.

        Args:
            p_value: P-value from statistical test

        Returns:
            Significance level string (e.g., '*** (p<0.001)')
        """
        if p_value < 0.001:
            return '*** (p<0.001)'
        elif p_value < 0.01:
            return '** (p<0.01)'
        elif p_value < 0.05:
            return '* (p<0.05)'
        else:
            return 'ns (p≥0.05)'

    def export_to_excel(self, tables: Dict[str, pd.DataFrame],
                        output_path: Path) -> None:
        """Export statistical tables to formatted Excel file.

        Args:
            tables: Dictionary of table name to DataFrame
            output_path: Path for output Excel file
        """
        wb = Workbook()
        wb.remove(wb.active)

        # Create worksheets for each table type
        for table_name, df in tables.items():
            ws = wb.create_sheet(table_name.capitalize())

            # Write header
            for col_idx, col_name in enumerate(df.columns, 1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.font = Font(bold=True, size=12, color='FFFFFF')
                cell.fill = PatternFill(
                    start_color='366092', end_color='366092', fill_type='solid'
                )
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # Write data
            for row_idx, row in enumerate(df.itertuples(index=False), 2):
                for col_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                    # Highlight significant results
                    if col_idx == len(row) and 'Significant' in df.columns:
                        if isinstance(value, str):
                            if value.startswith('***'):
                                cell.fill = PatternFill(
                                    start_color='FF6B6B', end_color='FF6B6B',
                                    fill_type='solid'
                                )
                            elif value.startswith('**'):
                                cell.fill = PatternFill(
                                    start_color='FFA07A', end_color='FFA07A',
                                    fill_type='solid'
                                )
                            elif value.startswith('*'):
                                cell.fill = PatternFill(
                                    start_color='FFD700', end_color='FFD700',
                                    fill_type='solid'
                                )

            # Auto-adjust column widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width

        wb.save(output_path)
