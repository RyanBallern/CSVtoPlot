"""Main GUI application for Neuromorpho Analyzer."""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from pathlib import Path
from typing import Optional, List, Dict
import threading
import pandas as pd

from .core.database import SQLiteDatabase
from .core.importers import UnifiedImporter, FileScanner, HeaderScanner, ParameterMapper
from .core.processors import StatisticsEngine, DensityCalculator, RepresentativeFileAnalyzer
from .core.exporters import ExcelExporter, ExportConfig, ExportParameterSelector, StatisticsTableExporter
from .core.profiles import AnalysisProfile, ProfileManager
from .gui.widgets import (
    ConditionNamesWidget,
    ColorPickerWidget,
    ExportConfigWidget,
    ConditionSelectorWidget,
    ParameterSelectorWidget,
)


class NeuromorphoAnalyzerApp:
    """Main application window for Neuromorpho Analyzer.

    Provides a complete GUI for:
    - Importing data from CSV/Excel/JSON files
    - Storing data in SQLite database
    - Running statistical analysis
    - Generating plots and exports
    - Managing analysis profiles
    """

    def __init__(self, root: tk.Tk):
        """Initialize the application.

        Args:
            root: Tkinter root window
        """
        self.root = root
        self.root.title("Neuromorpho Analyzer")
        self.root.geometry("1000x700")
        self.root.minsize(800, 600)

        # State
        self.database: Optional[SQLiteDatabase] = None
        self.current_assay_id: Optional[int] = None
        self.available_parameters: List[str] = []
        self.available_conditions: List[str] = []
        self.profile_manager = ProfileManager(Path.home() / ".neuromorpho_analyzer" / "profiles")

        # Create UI
        self._create_menu()
        self._create_main_layout()
        self._create_status_bar()

        # Initialize with empty state
        self._update_ui_state()

    def _create_menu(self):
        """Create application menu bar."""
        menubar = tk.Menu(self.root)
        self.root.config(menu=menubar)

        # File menu
        file_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="File", menu=file_menu)
        file_menu.add_command(label="New Database...", command=self._new_database)
        file_menu.add_command(label="Open Database...", command=self._open_database)
        file_menu.add_separator()
        file_menu.add_command(label="Import Files...", command=self._import_files)
        file_menu.add_command(label="Import Directory...", command=self._import_directory)
        file_menu.add_separator()
        file_menu.add_command(label="Exit", command=self.root.quit)

        # Analysis menu
        analysis_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Analysis", menu=analysis_menu)
        analysis_menu.add_command(label="Run Statistics", command=self._run_statistics)
        analysis_menu.add_command(label="Find Representative Files", command=self._find_representative)
        analysis_menu.add_command(label="Calculate Density", command=self._calculate_density)

        # Export menu
        export_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Export", menu=export_menu)
        export_menu.add_command(label="Export to Excel...", command=self._export_excel)
        export_menu.add_command(label="Export Statistics Tables...", command=self._export_statistics)

        # View/Preview menu
        view_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="View", menu=view_menu)
        view_menu.add_command(label="Preview Data Plot", command=self._preview_plot)
        view_menu.add_command(label="Preview Frequency Histogram", command=self._preview_histogram)
        view_menu.add_separator()
        view_menu.add_command(label="Open Excel File...", command=self._preview_xlsx)

        # Profiles menu
        profiles_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Profiles", menu=profiles_menu)
        profiles_menu.add_command(label="Save Profile...", command=self._save_profile)
        profiles_menu.add_command(label="Load Profile...", command=self._load_profile)

        # Help menu
        help_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Help", menu=help_menu)
        help_menu.add_command(label="About", command=self._show_about)

    def _create_main_layout(self):
        """Create main application layout."""
        # Main paned window
        self.paned = ttk.PanedWindow(self.root, orient='horizontal')
        self.paned.pack(fill='both', expand=True, padx=5, pady=5)

        # Left panel - Data & Parameters
        left_frame = ttk.Frame(self.paned, width=300)
        self.paned.add(left_frame, weight=1)

        # Database info
        db_frame = ttk.LabelFrame(left_frame, text="Database", padding=10)
        db_frame.pack(fill='x', padx=5, pady=5)

        self.db_label = ttk.Label(db_frame, text="No database loaded")
        self.db_label.pack(anchor='w')

        self.assay_label = ttk.Label(db_frame, text="")
        self.assay_label.pack(anchor='w')

        self.measurements_label = ttk.Label(db_frame, text="")
        self.measurements_label.pack(anchor='w')

        # Parameter selector
        param_frame = ttk.LabelFrame(left_frame, text="Parameters", padding=10)
        param_frame.pack(fill='both', expand=True, padx=5, pady=5)

        self.param_selector = ParameterSelectorWidget(
            param_frame,
            parameters=[],
            callback=self._on_parameter_change
        )
        self.param_selector.pack(fill='both', expand=True)

        # Right panel - Conditions & Export
        right_frame = ttk.Frame(self.paned, width=400)
        self.paned.add(right_frame, weight=2)

        # Notebook for tabs
        self.notebook = ttk.Notebook(right_frame)
        self.notebook.pack(fill='both', expand=True, padx=5, pady=5)

        # Conditions tab
        conditions_frame = ttk.Frame(self.notebook, padding=10)
        self.notebook.add(conditions_frame, text="Conditions")

        self.condition_selector = ConditionSelectorWidget(
            conditions_frame,
            conditions=[],
            callback=self._on_condition_change
        )
        self.condition_selector.pack(fill='x')

        ttk.Separator(conditions_frame, orient='horizontal').pack(fill='x', pady=10)

        self.condition_names = ConditionNamesWidget(
            conditions_frame,
            conditions=[],
            callback=self._on_condition_names_change
        )
        self.condition_names.pack(fill='x')

        # Export tab
        export_frame = ttk.Frame(self.notebook, padding=10)
        self.notebook.add(export_frame, text="Export")

        self.export_config = ExportConfigWidget(
            export_frame,
            callback=self._on_export_config_change
        )
        self.export_config.pack(fill='x')

        # Export button
        ttk.Button(
            export_frame,
            text="Export Now",
            command=self._export_excel
        ).pack(pady=20)

        # Results tab
        results_frame = ttk.Frame(self.notebook, padding=10)
        self.notebook.add(results_frame, text="Results")

        # Results text area
        self.results_text = tk.Text(results_frame, height=20, wrap='word')
        results_scrollbar = ttk.Scrollbar(results_frame, command=self.results_text.yview)
        self.results_text.configure(yscrollcommand=results_scrollbar.set)

        self.results_text.pack(side='left', fill='both', expand=True)
        results_scrollbar.pack(side='right', fill='y')

    def _create_status_bar(self):
        """Create status bar at bottom of window."""
        self.status_var = tk.StringVar(value="Ready")
        status_bar = ttk.Label(
            self.root,
            textvariable=self.status_var,
            relief='sunken',
            anchor='w'
        )
        status_bar.pack(side='bottom', fill='x')

    def _update_ui_state(self):
        """Update UI based on current state."""
        if self.database:
            self.db_label.config(text=f"Database: {self.database.db_path.name}")

            if self.current_assay_id:
                count = self.database.get_measurement_count(self.current_assay_id)
                conditions = self.database.get_conditions(self.current_assay_id)
                params = self.database.get_parameters(self.current_assay_id)

                self.assay_label.config(text=f"Assay ID: {self.current_assay_id}")
                self.measurements_label.config(text=f"Measurements: {count}")

                self.available_conditions = conditions
                self.available_parameters = params

                self.param_selector.update_parameters(params)
                self.condition_selector.update_conditions(conditions)
                self.condition_names.update_conditions(conditions)
        else:
            self.db_label.config(text="No database loaded")
            self.assay_label.config(text="")
            self.measurements_label.config(text="")

    def _set_status(self, message: str):
        """Update status bar message."""
        self.status_var.set(message)
        self.root.update_idletasks()

    def _log_result(self, message: str):
        """Add message to results text area."""
        self.results_text.insert('end', message + "\n")
        self.results_text.see('end')

    def _get_quick_paths(self) -> List[tuple]:
        """Get list of quick access paths (name, path)."""
        quick_paths = [("Home", str(Path.home()))]

        # Common directories
        docs = Path.home() / "Documents"
        if docs.exists():
            quick_paths.append(("Documents", str(docs)))

        desktop = Path.home() / "Desktop"
        if desktop.exists():
            quick_paths.append(("Desktop", str(desktop)))

        # Linux mount points
        for mount in ['/run/media', '/media', '/mnt']:
            mount_path = Path(mount)
            if mount_path.exists():
                try:
                    for subdir in mount_path.iterdir():
                        if subdir.is_dir():
                            # Check subdirectories for user mounts
                            for user_mount in subdir.iterdir():
                                if user_mount.is_dir():
                                    quick_paths.append((user_mount.name[:12], str(user_mount)))
                except PermissionError:
                    pass

        # Windows drives
        import platform
        if platform.system() == 'Windows':
            import string
            for letter in string.ascii_uppercase:
                drive = f"{letter}:\\"
                if Path(drive).exists():
                    quick_paths.append((f"{letter}:", drive))

        return quick_paths[:8]  # Max 8 quick paths

    def _create_path_dialog(self, title: str, mode: str = 'file',
                            default_name: str = '', extension: str = '',
                            filetypes: list = None) -> Optional[str]:
        """Create flexible path selection dialog.

        Args:
            title: Dialog title
            mode: 'file' for file selection, 'folder' for folder, 'save' for save file
            default_name: Default filename for save mode
            extension: Default extension for save mode
            filetypes: List of (description, pattern) tuples for file browser

        Returns:
            Selected path or None if cancelled
        """
        dialog = tk.Toplevel(self.root)
        dialog.title(title)
        dialog.geometry("650x280")
        dialog.transient(self.root)
        dialog.grab_set()

        result = {'path': None}

        # Path entry
        ttk.Label(dialog, text="Enter path or use buttons below:").pack(pady=(10, 5), padx=10, anchor='w')

        path_frame = ttk.Frame(dialog)
        path_frame.pack(fill='x', padx=10, pady=5)

        path_var = tk.StringVar()
        path_entry = ttk.Entry(path_frame, textvariable=path_var, width=55)
        path_entry.pack(side='left', fill='x', expand=True)

        def browse():
            if mode == 'folder':
                selected = filedialog.askdirectory()
            elif mode == 'save':
                selected = filedialog.asksaveasfilename(
                    defaultextension=extension,
                    filetypes=filetypes or [("All Files", "*.*")]
                )
            else:
                selected = filedialog.askopenfilename(
                    filetypes=filetypes or [("All Files", "*.*")]
                )
            if selected:
                path_var.set(selected)

        ttk.Button(path_frame, text="Browse...", command=browse).pack(side='left', padx=(5, 0))

        # Quick access buttons
        quick_frame = ttk.LabelFrame(dialog, text="Quick Access", padding=5)
        quick_frame.pack(fill='x', padx=10, pady=5)

        quick_paths = self._get_quick_paths()
        for i, (name, path) in enumerate(quick_paths):
            row = i // 4
            col = i % 4
            ttk.Button(quick_frame, text=name, width=12,
                       command=lambda p=path: path_var.set(p)).grid(row=row, column=col, padx=2, pady=2)

        # New folder/file creation
        create_frame = ttk.LabelFrame(dialog, text="Create New", padding=5)
        create_frame.pack(fill='x', padx=10, pady=5)

        ttk.Label(create_frame, text="Name:").grid(row=0, column=0, padx=5)
        new_name_var = tk.StringVar(value=default_name)
        ttk.Entry(create_frame, textvariable=new_name_var, width=25).grid(row=0, column=1, padx=5)

        def create_folder():
            base = path_var.get().strip()
            name = new_name_var.get().strip()
            if base and name:
                new_path = Path(base) / name
                new_path.mkdir(parents=True, exist_ok=True)
                path_var.set(str(new_path))

        def create_file():
            base = path_var.get().strip()
            name = new_name_var.get().strip()
            if base and name:
                if extension and not name.endswith(extension):
                    name += extension
                new_path = Path(base) / name
                path_var.set(str(new_path))

        ttk.Button(create_frame, text="Create Folder", command=create_folder).grid(row=0, column=2, padx=5)
        if mode == 'save':
            ttk.Button(create_frame, text="Set Filename", command=create_file).grid(row=0, column=3, padx=5)

        # OK/Cancel buttons
        def on_ok():
            path = path_var.get().strip()
            if path:
                result['path'] = path
            dialog.destroy()

        button_frame = ttk.Frame(dialog)
        button_frame.pack(pady=15)
        ttk.Button(button_frame, text="OK", command=on_ok, width=10).pack(side='left', padx=5)
        ttk.Button(button_frame, text="Cancel", command=dialog.destroy, width=10).pack(side='left', padx=5)

        path_entry.focus_set()
        dialog.wait_window()

        return result['path']

    # --- File operations ---

    def _new_database(self):
        """Create a new database."""
        filepath = self._create_path_dialog(
            "Create New Database",
            mode='save',
            default_name='analysis.db',
            extension='.db',
            filetypes=[("SQLite Database", "*.db"), ("All Files", "*.*")]
        )
        if filepath:
            # Ensure .db extension
            if not filepath.endswith('.db'):
                filepath += '.db'
            self.database = SQLiteDatabase(filepath)
            self.database.connect()
            self.current_assay_id = None
            self._update_ui_state()
            self._set_status(f"Created new database: {filepath}")

    def _open_database(self):
        """Open an existing database."""
        filepath = self._create_path_dialog(
            "Open Database",
            mode='file',
            filetypes=[("SQLite Database", "*.db"), ("All Files", "*.*")]
        )
        if filepath and Path(filepath).exists():
            self.database = SQLiteDatabase(filepath)
            self.database.connect()

            # Get first assay if exists
            assays = self.database.list_assays()
            if assays:
                self.current_assay_id = assays[0]['id']

            self._update_ui_state()
            self._set_status(f"Opened database: {filepath}")

    def _import_files(self):
        """Import data files."""
        if not self.database:
            messagebox.showwarning("No Database", "Please create or open a database first.")
            return

        # Create dialog for file selection with manual path entry
        dialog = tk.Toplevel(self.root)
        dialog.title("Import Files")
        dialog.geometry("650x350")
        dialog.transient(self.root)
        dialog.grab_set()

        ttk.Label(dialog, text="Enter file paths (one per line) or use Browse:").pack(pady=(10, 5), padx=10, anchor='w')

        # Text area for multiple paths
        text_frame = ttk.Frame(dialog)
        text_frame.pack(fill='both', expand=True, padx=10, pady=5)

        paths_text = tk.Text(text_frame, height=8, width=60)
        scrollbar = ttk.Scrollbar(text_frame, command=paths_text.yview)
        paths_text.configure(yscrollcommand=scrollbar.set)
        paths_text.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')

        # Quick access
        quick_frame = ttk.LabelFrame(dialog, text="Quick Access", padding=5)
        quick_frame.pack(fill='x', padx=10, pady=5)

        quick_paths = self._get_quick_paths()
        for i, (name, path) in enumerate(quick_paths[:6]):
            ttk.Button(quick_frame, text=name, width=10,
                       command=lambda p=path: paths_text.insert('end', p + '\n')).grid(row=0, column=i, padx=2)

        # Browse button
        button_row = ttk.Frame(dialog)
        button_row.pack(fill='x', padx=10, pady=5)

        def browse_files():
            files = filedialog.askopenfilenames(
                filetypes=[
                    ("All Supported", "*.csv *.xlsx *.xls *.json"),
                    ("CSV Files", "*.csv"),
                    ("Excel Files", "*.xlsx *.xls"),
                    ("JSON Files", "*.json"),
                ]
            )
            if files:
                for f in files:
                    paths_text.insert('end', f + '\n')

        ttk.Button(button_row, text="Browse Files...", command=browse_files).pack(side='left', padx=5)
        ttk.Button(button_row, text="Clear", command=lambda: paths_text.delete('1.0', 'end')).pack(side='left', padx=5)

        def do_import():
            text = paths_text.get('1.0', 'end').strip()
            if not text:
                messagebox.showwarning("No Files", "Please enter or select files to import.")
                return

            paths = [p.strip() for p in text.split('\n') if p.strip()]
            valid_paths = [p for p in paths if Path(p).exists() and Path(p).is_file()]

            if not valid_paths:
                messagebox.showwarning("No Valid Files", "No valid file paths found.")
                return

            dialog.destroy()
            self._import_files_list(valid_paths)

        action_frame = ttk.Frame(dialog)
        action_frame.pack(pady=10)
        ttk.Button(action_frame, text="Import", command=do_import).pack(side='left', padx=5)
        ttk.Button(action_frame, text="Cancel", command=dialog.destroy).pack(side='left', padx=5)

    def _import_directory(self):
        """Import all data files from a directory."""
        if not self.database:
            messagebox.showwarning("No Database", "Please create or open a database first.")
            return

        # Create dialog with entry field for paste and browse button
        dialog = tk.Toplevel(self.root)
        dialog.title("Import Directory")
        dialog.geometry("600x200")
        dialog.transient(self.root)
        dialog.grab_set()

        ttk.Label(dialog, text="Enter or paste directory path:").pack(pady=(10, 5), padx=10, anchor='w')

        path_frame = ttk.Frame(dialog)
        path_frame.pack(fill='x', padx=10, pady=5)

        path_var = tk.StringVar()
        path_entry = ttk.Entry(path_frame, textvariable=path_var, width=50)
        path_entry.pack(side='left', fill='x', expand=True)

        def browse():
            dirpath = filedialog.askdirectory()
            if dirpath:
                path_var.set(dirpath)

        ttk.Button(path_frame, text="Browse...", command=browse).pack(side='left', padx=(5, 0))

        # Quick drive/location selection
        drives_frame = ttk.LabelFrame(dialog, text="Quick Access", padding=5)
        drives_frame.pack(fill='x', padx=10, pady=5)

        def set_path(path):
            path_var.set(path)

        # Detect available drives/mount points
        quick_paths = []
        # Linux mount points
        for mount in ['/run/media', '/media', '/mnt']:
            mount_path = Path(mount)
            if mount_path.exists():
                for subdir in mount_path.iterdir():
                    if subdir.is_dir():
                        quick_paths.append((subdir.name[:10], str(subdir)))
        # Home directory
        quick_paths.insert(0, ("Home", str(Path.home())))
        # Common locations
        docs = Path.home() / "Documents"
        if docs.exists():
            quick_paths.insert(1, ("Documents", str(docs)))

        # Windows drives (if on Windows)
        import platform
        if platform.system() == 'Windows':
            import string
            for letter in string.ascii_uppercase:
                drive = f"{letter}:\\"
                if Path(drive).exists():
                    quick_paths.append((f"{letter}:", drive))

        # Create buttons for quick paths (max 6)
        for i, (name, path) in enumerate(quick_paths[:6]):
            ttk.Button(drives_frame, text=name, width=10,
                       command=lambda p=path: set_path(p)).grid(row=0, column=i, padx=2)

        def do_import():
            dirpath = path_var.get().strip()
            if not dirpath:
                messagebox.showwarning("No Path", "Please enter a directory path.")
                return

            dirpath = Path(dirpath)
            if not dirpath.exists():
                messagebox.showerror("Error", f"Directory does not exist:\n{dirpath}")
                return

            if not dirpath.is_dir():
                messagebox.showerror("Error", f"Path is not a directory:\n{dirpath}")
                return

            dialog.destroy()
            self._scan_and_import_directory(dirpath)

        button_frame = ttk.Frame(dialog)
        button_frame.pack(pady=20)
        ttk.Button(button_frame, text="Import", command=do_import).pack(side='left', padx=5)
        ttk.Button(button_frame, text="Cancel", command=dialog.destroy).pack(side='left', padx=5)

        path_entry.focus_set()

    def _scan_and_import_directory(self, dirpath: Path):
        """Scan directory and import all supported files."""
        # First try FileScanner for structured naming
        scanner = FileScanner(dirpath)
        files = scanner.scan_files()

        # If no structured files found, scan for all supported files
        if not files:
            supported_extensions = {'.csv', '.xlsx', '.xls', '.json'}
            all_files = []
            for ext in supported_extensions:
                all_files.extend(dirpath.glob(f'*{ext}'))

            if all_files:
                # Show filename structure analysis dialog
                self._show_filename_structure_dialog(sorted(all_files))
                return

        if files:
            # Show condition assignment dialog
            self._show_condition_assignment_dialog(files)
        else:
            messagebox.showinfo("No Files", "No supported files found in directory.\n\nSupported formats: .csv, .xlsx, .xls, .json")

    def _show_filename_structure_dialog(self, all_files: List[Path]):
        """Analyze filename structure and let user choose condition substring."""
        dialog = tk.Toplevel(self.root)
        dialog.title("Filename Structure Analysis")
        dialog.geometry("800x500")
        dialog.transient(self.root)
        dialog.grab_set()

        # Analyze filename structure
        sample_files = all_files[:10]  # Use first 10 files as samples

        # Parse filenames into parts
        file_parts = []
        max_parts = 0
        for filepath in sample_files:
            stem = filepath.stem
            # Split by common delimiters
            parts = stem.replace('-', '_').replace('.', '_').split('_')
            file_parts.append((filepath.name, parts))
            max_parts = max(max_parts, len(parts))

        ttk.Label(dialog, text="Filename Structure Analysis",
                  font=('TkDefaultFont', 12, 'bold')).pack(pady=10)

        ttk.Label(dialog, text="Sample filenames split by '_' delimiter:").pack(anchor='w', padx=10)

        # Show filename structure table
        table_frame = ttk.Frame(dialog)
        table_frame.pack(fill='both', expand=True, padx=10, pady=5)

        # Create treeview
        columns = ['Filename'] + [f'Part {i+1}' for i in range(max_parts)]
        tree = ttk.Treeview(table_frame, columns=columns, show='headings', height=8)

        for col in columns:
            tree.heading(col, text=col)
            tree.column(col, width=100 if col != 'Filename' else 200)

        vsb = ttk.Scrollbar(table_frame, orient='vertical', command=tree.yview)
        hsb = ttk.Scrollbar(table_frame, orient='horizontal', command=tree.xview)
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        tree.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')
        hsb.grid(row=1, column=0, sticky='ew')
        table_frame.grid_rowconfigure(0, weight=1)
        table_frame.grid_columnconfigure(0, weight=1)

        # Populate table
        for filename, parts in file_parts:
            row = [filename] + parts + [''] * (max_parts - len(parts))
            tree.insert('', 'end', values=row)

        # Show unique values per part position
        ttk.Label(dialog, text="\nUnique values found in each position:",
                  font=('TkDefaultFont', 10, 'bold')).pack(anchor='w', padx=10, pady=(10, 5))

        unique_frame = ttk.Frame(dialog)
        unique_frame.pack(fill='x', padx=10)

        part_uniques = []
        for i in range(max_parts):
            values = set()
            for _, parts in file_parts:
                if i < len(parts):
                    values.add(parts[i])
            part_uniques.append(sorted(values))

            # Show unique values (truncated)
            values_str = ', '.join(list(values)[:5])
            if len(values) > 5:
                values_str += f'... ({len(values)} unique)'
            ttk.Label(unique_frame, text=f"Part {i+1}: {values_str}").pack(anchor='w')

        # Selection frame
        select_frame = ttk.LabelFrame(dialog, text="Select Condition Position", padding=10)
        select_frame.pack(fill='x', padx=10, pady=10)

        ttk.Label(select_frame, text="Which part contains the CONDITION?").pack(anchor='w')

        condition_pos_var = tk.IntVar(value=1)

        # Radio buttons for each position
        radio_frame = ttk.Frame(select_frame)
        radio_frame.pack(fill='x', pady=5)

        for i in range(max_parts):
            # Find most descriptive sample value
            sample_val = part_uniques[i][0] if part_uniques[i] else "?"
            ttk.Radiobutton(
                radio_frame,
                text=f"Part {i+1} (e.g., '{sample_val}')",
                variable=condition_pos_var,
                value=i
            ).pack(side='left', padx=10)

        # Custom delimiter option
        delim_frame = ttk.Frame(select_frame)
        delim_frame.pack(fill='x', pady=5)

        ttk.Label(delim_frame, text="Or use custom delimiter:").pack(side='left')
        custom_delim_var = tk.StringVar(value='_')
        ttk.Entry(delim_frame, textvariable=custom_delim_var, width=5).pack(side='left', padx=5)

        def apply_selection():
            condition_pos = condition_pos_var.get()
            delimiter = custom_delim_var.get() or '_'

            # Create file info with selected condition position
            files = []
            for filepath in all_files:
                stem = filepath.stem
                parts = stem.replace('-', delimiter).split(delimiter)

                if condition_pos < len(parts):
                    condition = parts[condition_pos]
                else:
                    condition = stem

                files.append({
                    'path': filepath,
                    'condition': condition,
                    'file_format': filepath.suffix[1:]
                })

            dialog.destroy()
            self._show_condition_assignment_dialog(files)

        # Buttons
        button_frame = ttk.Frame(dialog)
        button_frame.pack(pady=10)
        ttk.Button(button_frame, text="Apply & Continue", command=apply_selection).pack(side='left', padx=5)
        ttk.Button(button_frame, text="Cancel", command=dialog.destroy).pack(side='left', padx=5)

    def _show_condition_assignment_dialog(self, files: List[Dict]):
        """Show dialog to assign conditions to files."""
        dialog = tk.Toplevel(self.root)
        dialog.title("Assign Conditions")
        dialog.geometry("700x500")
        dialog.transient(self.root)
        dialog.grab_set()

        ttk.Label(dialog, text="Review and edit conditions for each file:",
                  font=('TkDefaultFont', 10, 'bold')).pack(pady=10, padx=10, anchor='w')

        # Frame with scrollbar for file list
        list_frame = ttk.Frame(dialog)
        list_frame.pack(fill='both', expand=True, padx=10, pady=5)

        canvas = tk.Canvas(list_frame)
        scrollbar = ttk.Scrollbar(list_frame, orient='vertical', command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)

        scrollable_frame.bind(
            '<Configure>',
            lambda e: canvas.configure(scrollregion=canvas.bbox('all'))
        )

        canvas.create_window((0, 0), window=scrollable_frame, anchor='nw')
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')

        # Headers
        ttk.Label(scrollable_frame, text="Filename", font=('TkDefaultFont', 9, 'bold'), width=40).grid(row=0, column=0, padx=5, pady=2, sticky='w')
        ttk.Label(scrollable_frame, text="Condition", font=('TkDefaultFont', 9, 'bold'), width=20).grid(row=0, column=1, padx=5, pady=2, sticky='w')
        ttk.Label(scrollable_frame, text="Dataset (L/T)", font=('TkDefaultFont', 9, 'bold'), width=10).grid(row=0, column=2, padx=5, pady=2, sticky='w')

        # Store entry widgets
        condition_vars = []
        dataset_vars = []

        for i, file_info in enumerate(files):
            filename = file_info['path'].name

            # Filename label
            ttk.Label(scrollable_frame, text=filename, width=40).grid(row=i+1, column=0, padx=5, pady=2, sticky='w')

            # Condition entry
            cond_var = tk.StringVar(value=file_info.get('condition', 'Unknown'))
            cond_entry = ttk.Entry(scrollable_frame, textvariable=cond_var, width=20)
            cond_entry.grid(row=i+1, column=1, padx=5, pady=2)
            condition_vars.append(cond_var)

            # Dataset marker dropdown (L/T/None)
            ds_var = tk.StringVar(value=file_info.get('dataset_marker', ''))
            ds_combo = ttk.Combobox(scrollable_frame, textvariable=ds_var, values=['', 'L', 'T'], width=8, state='readonly')
            ds_combo.grid(row=i+1, column=2, padx=5, pady=2)
            dataset_vars.append(ds_var)

        # Bulk assignment frame
        bulk_frame = ttk.LabelFrame(dialog, text="Bulk Assign", padding=10)
        bulk_frame.pack(fill='x', padx=10, pady=10)

        ttk.Label(bulk_frame, text="Set all conditions to:").grid(row=0, column=0, padx=5)
        bulk_cond_var = tk.StringVar()
        bulk_cond_entry = ttk.Entry(bulk_frame, textvariable=bulk_cond_var, width=15)
        bulk_cond_entry.grid(row=0, column=1, padx=5)

        def apply_bulk_condition():
            val = bulk_cond_var.get().strip()
            if val:
                for var in condition_vars:
                    var.set(val)

        ttk.Button(bulk_frame, text="Apply", command=apply_bulk_condition).grid(row=0, column=2, padx=5)

        ttk.Label(bulk_frame, text="Set all datasets to:").grid(row=0, column=3, padx=(20, 5))
        bulk_ds_var = tk.StringVar()
        bulk_ds_combo = ttk.Combobox(bulk_frame, textvariable=bulk_ds_var, values=['', 'L', 'T'], width=5, state='readonly')
        bulk_ds_combo.grid(row=0, column=4, padx=5)

        def apply_bulk_dataset():
            val = bulk_ds_var.get()
            for var in dataset_vars:
                var.set(val)

        ttk.Button(bulk_frame, text="Apply", command=apply_bulk_dataset).grid(row=0, column=5, padx=5)

        # Import button
        def do_import():
            # Update file info with user-assigned conditions
            for i, file_info in enumerate(files):
                file_info['condition'] = condition_vars[i].get().strip() or 'Unknown'
                file_info['dataset_marker'] = dataset_vars[i].get() or None

            dialog.destroy()
            filepaths = [str(f['path']) for f in files]
            self._import_files_list(filepaths, files)

        button_frame = ttk.Frame(dialog)
        button_frame.pack(pady=10)
        ttk.Button(button_frame, text="Import", command=do_import).pack(side='left', padx=5)
        ttk.Button(button_frame, text="Cancel", command=dialog.destroy).pack(side='left', padx=5)

    def _import_files_list(self, filepaths: List[str], file_infos: List[Dict] = None):
        """Import a list of files."""
        self._set_status("Importing files...")

        # Create new assay
        assay_name = f"Import {Path(filepaths[0]).parent.name}"
        self.current_assay_id = self.database.insert_assay(assay_name)

        total_imported = 0
        for i, filepath in enumerate(filepaths):
            try:
                df = UnifiedImporter.import_file(filepath)

                # Get condition from file info or filename
                condition = "Unknown"
                if file_infos and i < len(file_infos):
                    condition = file_infos[i].get('condition', 'Unknown')
                else:
                    # Try to parse from filename
                    parts = Path(filepath).stem.split('_')
                    if len(parts) >= 2:
                        condition = parts[1]

                count = self.database.insert_measurements(
                    self.current_assay_id,
                    df,
                    source_file=Path(filepath).name,
                    condition=condition
                )
                total_imported += count
                self._set_status(f"Imported {i+1}/{len(filepaths)}: {Path(filepath).name}")

            except Exception as e:
                self._log_result(f"Error importing {filepath}: {e}")

        self._update_ui_state()
        self._set_status(f"Imported {total_imported} measurements from {len(filepaths)} files")
        self._log_result(f"Successfully imported {total_imported} measurements")

    # --- Analysis operations ---

    def _run_statistics(self):
        """Run statistical analysis on current data."""
        if not self._check_data_loaded():
            return

        self._set_status("Running statistics...")
        self.notebook.select(2)  # Switch to results tab

        try:
            stats = StatisticsEngine()
            df = self.database.get_measurements(self.current_assay_id)

            selected_params = self.param_selector.get_selected_parameters()
            selected_conditions = self.condition_selector.get_selected_conditions()

            # Filter by conditions
            df = df[df['condition'].isin(selected_conditions)]

            # Convert wide to long format
            df = self._wide_to_long(df, selected_params)

            self._log_result("=" * 50)
            self._log_result("STATISTICAL ANALYSIS")
            self._log_result("=" * 50)

            for param in selected_params:
                param_df = df[df['parameter_name'] == param]
                if param_df.empty:
                    continue

                self._log_result(f"\nParameter: {param}")
                self._log_result("-" * 30)

                result = stats.auto_compare(param_df, 'value', 'condition')

                main_test = result.get('main_test')
                if main_test:
                    self._log_result(f"Test: {main_test.test_name}")
                    self._log_result(f"Statistic: {main_test.statistic:.4f}")
                    self._log_result(f"P-value: {main_test.p_value:.4e}")
                    self._log_result(f"Significant: {'Yes' if main_test.significant else 'No'}")

            self._set_status("Statistics complete")

        except Exception as e:
            self._log_result(f"Error: {e}")
            self._set_status("Statistics failed")

    def _find_representative(self):
        """Find representative files per condition."""
        if not self._check_data_loaded():
            return

        self._set_status("Finding representative files...")
        self.notebook.select(2)

        try:
            analyzer = RepresentativeFileAnalyzer(self.database)
            selected_params = self.param_selector.get_selected_parameters()

            results = analyzer.analyze([self.current_assay_id], selected_params)

            self._log_result("=" * 50)
            self._log_result("REPRESENTATIVE FILES")
            self._log_result("=" * 50)

            for condition in results['condition'].unique():
                self._log_result(f"\n{condition}:")
                cond_results = results[results['condition'] == condition].head(3)
                for _, row in cond_results.iterrows():
                    self._log_result(f"  {row['rank']}. {row['file']} (dist: {row['distance_from_average']:.4f})")

            self._set_status("Representative analysis complete")

        except Exception as e:
            self._log_result(f"Error: {e}")
            self._set_status("Analysis failed")

    def _calculate_density(self):
        """Calculate structure density."""
        if not self._check_data_loaded():
            return

        self._set_status("Calculating density...")
        self.notebook.select(2)

        try:
            calc = DensityCalculator()
            df = self.database.get_measurements(self.current_assay_id)

            self._log_result("=" * 50)
            self._log_result("DENSITY ANALYSIS")
            self._log_result(f"Image area: {calc.config.image_area:.4f} µm² (3.5021²)")
            self._log_result("=" * 50)

            for condition in self.condition_selector.get_selected_conditions():
                cond_df = df[df['condition'] == condition]
                count = len(cond_df)
                result = calc.calculate_density_from_count(count, condition=condition)

                self._log_result(f"\n{condition}:")
                self._log_result(f"  Count: {result.count}")
                self._log_result(f"  Density: {result.density_per_mm2:.2f} /mm²")

            self._set_status("Density calculation complete")

        except Exception as e:
            self._log_result(f"Error: {e}")
            self._set_status("Calculation failed")

    # --- Export operations ---

    def _export_excel(self):
        """Export data to Excel with full analysis."""
        if not self._check_data_loaded():
            return

        # Create export dialog with path entry and folder creation
        dialog = tk.Toplevel(self.root)
        dialog.title("Export to Excel")
        dialog.geometry("600x250")
        dialog.transient(self.root)
        dialog.grab_set()

        ttk.Label(dialog, text="Select or create output folder:").pack(pady=(10, 5), padx=10, anchor='w')

        path_frame = ttk.Frame(dialog)
        path_frame.pack(fill='x', padx=10, pady=5)

        path_var = tk.StringVar(value=str(Path.home() / "Documents"))
        path_entry = ttk.Entry(path_frame, textvariable=path_var, width=50)
        path_entry.pack(side='left', fill='x', expand=True)

        def browse():
            dirpath = filedialog.askdirectory()
            if dirpath:
                path_var.set(dirpath)

        ttk.Button(path_frame, text="Browse...", command=browse).pack(side='left', padx=(5, 0))

        # Quick drive/location selection
        drives_frame = ttk.LabelFrame(dialog, text="Quick Access", padding=5)
        drives_frame.pack(fill='x', padx=10, pady=5)

        def set_path(path):
            path_var.set(path)

        # Detect available drives/mount points
        quick_paths = [("Home", str(Path.home()))]
        docs = Path.home() / "Documents"
        if docs.exists():
            quick_paths.append(("Documents", str(docs)))

        # Linux mount points
        for mount in ['/run/media', '/media', '/mnt']:
            mount_path = Path(mount)
            if mount_path.exists():
                for subdir in mount_path.iterdir():
                    if subdir.is_dir():
                        quick_paths.append((subdir.name[:10], str(subdir)))

        # Windows drives
        import platform
        if platform.system() == 'Windows':
            import string
            for letter in string.ascii_uppercase:
                drive = f"{letter}:\\"
                if Path(drive).exists():
                    quick_paths.append((f"{letter}:", drive))

        for i, (name, path) in enumerate(quick_paths[:6]):
            ttk.Button(drives_frame, text=name, width=10,
                       command=lambda p=path: set_path(p)).grid(row=0, column=i, padx=2)

        # New folder creation
        newfolder_frame = ttk.Frame(dialog)
        newfolder_frame.pack(fill='x', padx=10, pady=5)

        ttk.Label(newfolder_frame, text="New folder name:").pack(side='left')
        newfolder_var = tk.StringVar()
        ttk.Entry(newfolder_frame, textvariable=newfolder_var, width=20).pack(side='left', padx=5)

        def create_folder():
            base = path_var.get().strip()
            name = newfolder_var.get().strip()
            if base and name:
                new_path = Path(base) / name
                new_path.mkdir(parents=True, exist_ok=True)
                path_var.set(str(new_path))
                newfolder_var.set("")

        ttk.Button(newfolder_frame, text="Create", command=create_folder).pack(side='left', padx=5)

        def do_export():
            output_dir = path_var.get().strip()
            if not output_dir:
                messagebox.showwarning("No Path", "Please select an output directory.")
                return

            output_path = Path(output_dir)
            output_path.mkdir(parents=True, exist_ok=True)

            dialog.destroy()
            self._perform_export(output_path)

        button_frame = ttk.Frame(dialog)
        button_frame.pack(pady=15)
        ttk.Button(button_frame, text="Export", command=do_export).pack(side='left', padx=5)
        ttk.Button(button_frame, text="Cancel", command=dialog.destroy).pack(side='left', padx=5)

    def _perform_export(self, output_dir: Path):
        """Perform the actual export with analysis."""
        self._set_status("Running analysis and exporting...")
        self.notebook.select(2)  # Switch to results tab

        try:
            # Run statistics first
            self._log_result("=" * 50)
            self._log_result("RUNNING COMPLETE ANALYSIS & EXPORT")
            self._log_result("=" * 50)

            stats = StatisticsEngine()
            param_selector = ExportParameterSelector(self.database)
            selected_params = self.param_selector.get_selected_parameters()
            param_selector.select_parameters(selected_params)

            # Log analysis summary
            df = self.database.get_measurements(self.current_assay_id)
            df_long = self._wide_to_long(df, selected_params)

            for param in selected_params:
                param_df = df_long[df_long['parameter_name'] == param]
                if param_df.empty:
                    continue

                self._log_result(f"\nParameter: {param}")
                self._log_result("-" * 30)

                try:
                    result = stats.auto_compare(param_df, 'value', 'condition')
                    main_test = result.get('main_test')
                    if main_test:
                        self._log_result(f"Test: {main_test.test_name}")
                        self._log_result(f"P-value: {main_test.p_value:.4e}")
                        self._log_result(f"Significant: {'Yes' if main_test.significant else 'No'}")
                except Exception as e:
                    self._log_result(f"Stats error: {e}")

            # Export to Excel
            self._log_result("\nExporting to Excel...")
            exporter = ExcelExporter(param_selector, stats)
            output_path = exporter.export(
                [self.current_assay_id],
                output_dir,
                self.database
            )

            self._set_status(f"Exported to: {output_path}")
            self._log_result(f"\nExport complete: {output_path}")
            messagebox.showinfo("Export Complete", f"Analysis and export saved to:\n{output_path}")

        except Exception as e:
            import traceback
            self._log_result(f"Export error: {e}")
            self._log_result(traceback.format_exc())
            self._set_status("Export failed")
            messagebox.showerror("Export Error", str(e))

    def _export_statistics(self):
        """Export statistics tables."""
        if not self._check_data_loaded():
            return

        filepath = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")]
        )
        if not filepath:
            return

        self._set_status("Exporting statistics...")

        try:
            stats = StatisticsEngine()
            exporter = StatisticsTableExporter(stats)

            df = self.database.get_measurements(self.current_assay_id)
            selected_params = self.param_selector.get_selected_parameters()

            # Build data dict for first parameter
            if selected_params:
                param = selected_params[0]
                param_df = df[df['parameter_name'] == param]
                data_dict = {}
                for cond in param_df['condition'].unique():
                    data_dict[cond] = param_df[param_df['condition'] == cond]['value']

                tables = exporter.create_statistics_tables(data_dict, param)
                exporter.export_to_excel(tables, Path(filepath))

                self._set_status(f"Exported statistics to: {filepath}")
                messagebox.showinfo("Export Complete", f"Saved to:\n{filepath}")

        except Exception as e:
            self._log_result(f"Export error: {e}")
            self._set_status("Export failed")
            messagebox.showerror("Export Error", str(e))

    # --- Profile operations ---

    def _save_profile(self):
        """Save current settings as profile."""
        name = tk.simpledialog.askstring("Save Profile", "Profile name:")
        if not name:
            return

        try:
            profile = AnalysisProfile(
                name=name,
                description="Saved from GUI",
                import_parameters=self.param_selector.get_selected_parameters(),
                selected_conditions=self.condition_selector.get_selected_conditions(),
            )
            self.profile_manager.save_profile(profile)
            self._set_status(f"Profile saved: {name}")
            messagebox.showinfo("Profile Saved", f"Profile '{name}' saved successfully.")

        except Exception as e:
            messagebox.showerror("Error", str(e))

    def _load_profile(self):
        """Load a saved profile."""
        profiles = self.profile_manager.list_profiles()
        if not profiles:
            messagebox.showinfo("No Profiles", "No saved profiles found.")
            return

        # Simple selection dialog
        dialog = tk.Toplevel(self.root)
        dialog.title("Load Profile")
        dialog.geometry("300x200")

        ttk.Label(dialog, text="Select profile:").pack(pady=10)

        listbox = tk.Listbox(dialog)
        for p in profiles:
            listbox.insert('end', p)
        listbox.pack(fill='both', expand=True, padx=10)

        def on_select():
            selection = listbox.curselection()
            if selection:
                name = profiles[selection[0]]
                profile = self.profile_manager.load_profile(name)

                if profile.import_parameters:
                    self.param_selector.set_selected(profile.import_parameters)
                if profile.selected_conditions:
                    self.condition_selector.set_selected(profile.selected_conditions)

                self._set_status(f"Profile loaded: {name}")
                dialog.destroy()

        ttk.Button(dialog, text="Load", command=on_select).pack(pady=10)

    # --- Callbacks ---

    def _on_parameter_change(self):
        """Called when parameter selection changes."""
        pass

    def _on_condition_change(self):
        """Called when condition selection changes."""
        pass

    def _on_condition_names_change(self):
        """Called when condition names change."""
        pass

    def _on_export_config_change(self):
        """Called when export config changes."""
        pass

    # --- Helpers ---

    def _wide_to_long(self, df: pd.DataFrame, parameters: List[str] = None) -> pd.DataFrame:
        """Convert wide format DataFrame to long format.

        Wide format: each parameter is a column
        Long format: parameter_name, value columns

        Args:
            df: Wide format DataFrame with source_file, condition columns
            parameters: List of parameter columns to melt (None = auto-detect)

        Returns:
            Long format DataFrame with parameter_name, value, source_file, condition
        """
        if df.empty:
            return pd.DataFrame(columns=['parameter_name', 'value', 'source_file', 'condition'])

        # Identify parameter columns (exclude metadata)
        metadata_cols = {'source_file', 'condition', 'assay_id', 'id', 'created_at'}
        if parameters is None:
            parameters = [col for col in df.columns if col not in metadata_cols]

        if not parameters:
            return pd.DataFrame(columns=['parameter_name', 'value', 'source_file', 'condition'])

        # Melt wide to long
        id_vars = [col for col in ['source_file', 'condition'] if col in df.columns]
        long_df = df.melt(
            id_vars=id_vars,
            value_vars=parameters,
            var_name='parameter_name',
            value_name='value'
        )

        return long_df

    def _check_data_loaded(self) -> bool:
        """Check if data is loaded and show warning if not."""
        if not self.database or not self.current_assay_id:
            messagebox.showwarning("No Data", "Please load data first.")
            return False
        return True

    # --- Preview operations ---

    def _preview_plot(self):
        """Preview data as bar plot with error bars."""
        if not self._check_data_loaded():
            return

        try:
            import matplotlib.pyplot as plt
            from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

            df = self.database.get_measurements(self.current_assay_id)
            selected_params = self.param_selector.get_selected_parameters()
            selected_conditions = self.condition_selector.get_selected_conditions()

            if not selected_params:
                messagebox.showwarning("No Parameters", "Please select at least one parameter.")
                return

            # Convert to long format
            df_long = self._wide_to_long(df, selected_params)
            df_long = df_long[df_long['condition'].isin(selected_conditions)]

            # Create preview window
            preview = tk.Toplevel(self.root)
            preview.title("Data Preview - Bar Plot")
            preview.geometry("900x700")

            # Create notebook for multiple parameters
            notebook = ttk.Notebook(preview)
            notebook.pack(fill='both', expand=True, padx=5, pady=5)

            for param in selected_params[:5]:  # Limit to 5 params
                param_df = df_long[df_long['parameter_name'] == param]
                if param_df.empty:
                    continue

                # Create figure
                fig, ax = plt.subplots(figsize=(8, 5))

                # Calculate means and SEMs
                stats_data = param_df.groupby('condition')['value'].agg(['mean', 'sem', 'count']).reindex(selected_conditions)

                x = range(len(selected_conditions))
                bars = ax.bar(x, stats_data['mean'], yerr=stats_data['sem'],
                             capsize=5, color='steelblue', edgecolor='black')

                ax.set_xticks(x)
                ax.set_xticklabels(selected_conditions, rotation=45, ha='right')
                ax.set_ylabel(param)
                ax.set_title(f'{param} by Condition')
                ax.spines['top'].set_visible(False)
                ax.spines['right'].set_visible(False)

                fig.tight_layout()

                # Add to notebook
                frame = ttk.Frame(notebook)
                notebook.add(frame, text=param[:15])

                canvas = FigureCanvasTkAgg(fig, master=frame)
                canvas.draw()
                canvas.get_tk_widget().pack(fill='both', expand=True)

            # Save button
            def save_plots():
                save_dir = filedialog.askdirectory(title="Select folder to save plots")
                if save_dir:
                    for param in selected_params:
                        param_df = df_long[df_long['parameter_name'] == param]
                        if param_df.empty:
                            continue

                        fig, ax = plt.subplots(figsize=(8, 5))
                        stats_data = param_df.groupby('condition')['value'].agg(['mean', 'sem']).reindex(selected_conditions)
                        x = range(len(selected_conditions))
                        ax.bar(x, stats_data['mean'], yerr=stats_data['sem'],
                               capsize=5, color='steelblue', edgecolor='black')
                        ax.set_xticks(x)
                        ax.set_xticklabels(selected_conditions, rotation=45, ha='right')
                        ax.set_ylabel(param)
                        ax.set_title(f'{param} by Condition')
                        ax.spines['top'].set_visible(False)
                        ax.spines['right'].set_visible(False)
                        fig.tight_layout()

                        clean_param = param.replace('/', '_').replace('\\', '_')
                        fig.savefig(Path(save_dir) / f'{clean_param}_barplot.png', dpi=150)
                        plt.close(fig)

                    messagebox.showinfo("Saved", f"Plots saved to {save_dir}")

            ttk.Button(preview, text="Save All Plots", command=save_plots).pack(pady=10)

        except ImportError as e:
            messagebox.showerror("Error", f"Matplotlib required for plots: {e}")
        except Exception as e:
            messagebox.showerror("Error", f"Plot error: {e}")

    def _preview_histogram(self):
        """Preview frequency histogram."""
        if not self._check_data_loaded():
            return

        try:
            import matplotlib.pyplot as plt
            from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

            df = self.database.get_measurements(self.current_assay_id)
            selected_params = self.param_selector.get_selected_parameters()
            selected_conditions = self.condition_selector.get_selected_conditions()

            if not selected_params:
                messagebox.showwarning("No Parameters", "Please select at least one parameter.")
                return

            # Convert to long format
            df_long = self._wide_to_long(df, selected_params)
            df_long = df_long[df_long['condition'].isin(selected_conditions)]

            # Create preview window
            preview = tk.Toplevel(self.root)
            preview.title("Data Preview - Histogram")
            preview.geometry("900x700")

            notebook = ttk.Notebook(preview)
            notebook.pack(fill='both', expand=True, padx=5, pady=5)

            for param in selected_params[:5]:
                param_df = df_long[df_long['parameter_name'] == param]
                if param_df.empty:
                    continue

                fig, ax = plt.subplots(figsize=(8, 5))

                # Plot histogram for each condition
                for cond in selected_conditions:
                    cond_values = param_df[param_df['condition'] == cond]['value'].dropna()
                    if len(cond_values) > 0:
                        ax.hist(cond_values, bins=30, alpha=0.6, label=cond, edgecolor='black')

                ax.set_xlabel(param)
                ax.set_ylabel('Frequency')
                ax.set_title(f'{param} Distribution')
                ax.legend()
                ax.spines['top'].set_visible(False)
                ax.spines['right'].set_visible(False)

                fig.tight_layout()

                frame = ttk.Frame(notebook)
                notebook.add(frame, text=param[:15])

                canvas = FigureCanvasTkAgg(fig, master=frame)
                canvas.draw()
                canvas.get_tk_widget().pack(fill='both', expand=True)

            # Save button
            def save_histograms():
                save_dir = filedialog.askdirectory(title="Select folder to save histograms")
                if save_dir:
                    for param in selected_params:
                        param_df = df_long[df_long['parameter_name'] == param]
                        if param_df.empty:
                            continue

                        fig, ax = plt.subplots(figsize=(8, 5))
                        for cond in selected_conditions:
                            cond_values = param_df[param_df['condition'] == cond]['value'].dropna()
                            if len(cond_values) > 0:
                                ax.hist(cond_values, bins=30, alpha=0.6, label=cond, edgecolor='black')
                        ax.set_xlabel(param)
                        ax.set_ylabel('Frequency')
                        ax.set_title(f'{param} Distribution')
                        ax.legend()
                        ax.spines['top'].set_visible(False)
                        ax.spines['right'].set_visible(False)
                        fig.tight_layout()

                        clean_param = param.replace('/', '_').replace('\\', '_')
                        fig.savefig(Path(save_dir) / f'{clean_param}_histogram.png', dpi=150)
                        plt.close(fig)

                    messagebox.showinfo("Saved", f"Histograms saved to {save_dir}")

            ttk.Button(preview, text="Save All Histograms", command=save_histograms).pack(pady=10)

        except ImportError as e:
            messagebox.showerror("Error", f"Matplotlib required for plots: {e}")
        except Exception as e:
            messagebox.showerror("Error", f"Plot error: {e}")

    def _preview_xlsx(self):
        """Preview an Excel file."""
        filepath = self._create_path_dialog(
            "Open Excel File",
            mode='file',
            filetypes=[("Excel Files", "*.xlsx *.xls"), ("All Files", "*.*")]
        )

        if not filepath or not Path(filepath).exists():
            return

        try:
            from openpyxl import load_workbook

            wb = load_workbook(filepath, read_only=True, data_only=True)

            # Create preview window
            preview = tk.Toplevel(self.root)
            preview.title(f"Excel Preview - {Path(filepath).name}")
            preview.geometry("1000x600")

            # Notebook for sheets
            notebook = ttk.Notebook(preview)
            notebook.pack(fill='both', expand=True, padx=5, pady=5)

            for sheet_name in wb.sheetnames[:10]:  # Limit to 10 sheets
                sheet = wb[sheet_name]

                frame = ttk.Frame(notebook)
                notebook.add(frame, text=sheet_name[:15])

                # Create treeview for data
                tree_frame = ttk.Frame(frame)
                tree_frame.pack(fill='both', expand=True)

                # Get data from sheet
                data = []
                max_cols = 0
                for row in sheet.iter_rows(max_row=100):  # Limit rows
                    row_data = [cell.value for cell in row]
                    max_cols = max(max_cols, len(row_data))
                    data.append(row_data)

                if not data:
                    ttk.Label(frame, text="(Empty sheet)").pack()
                    continue

                # Create treeview
                columns = [f'Col{i+1}' for i in range(max_cols)]
                tree = ttk.Treeview(tree_frame, columns=columns, show='headings')

                # Scrollbars
                vsb = ttk.Scrollbar(tree_frame, orient='vertical', command=tree.yview)
                hsb = ttk.Scrollbar(tree_frame, orient='horizontal', command=tree.xview)
                tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

                tree.grid(row=0, column=0, sticky='nsew')
                vsb.grid(row=0, column=1, sticky='ns')
                hsb.grid(row=1, column=0, sticky='ew')
                tree_frame.grid_rowconfigure(0, weight=1)
                tree_frame.grid_columnconfigure(0, weight=1)

                # Set column headers (first row if it looks like headers)
                for i, col in enumerate(columns):
                    header = data[0][i] if data and i < len(data[0]) and data[0][i] else f'Col {i+1}'
                    tree.heading(col, text=str(header)[:20])
                    tree.column(col, width=100)

                # Add data rows (skip first if used as header)
                start_row = 1 if data and any(isinstance(v, str) for v in data[0] if v) else 0
                for row in data[start_row:]:
                    # Pad row to max_cols
                    padded = row + [None] * (max_cols - len(row))
                    display = [str(v)[:50] if v is not None else '' for v in padded]
                    tree.insert('', 'end', values=display)

            wb.close()

            # Info label
            ttk.Label(preview, text=f"Sheets: {len(wb.sheetnames)} | Showing first 100 rows per sheet").pack(pady=5)

        except Exception as e:
            messagebox.showerror("Error", f"Failed to open Excel file: {e}")

    def _show_about(self):
        """Show about dialog."""
        messagebox.showinfo(
            "About Neuromorpho Analyzer",
            "Neuromorpho Analyzer v0.1.0\n\n"
            "A tool for neuromorphological data analysis.\n\n"
            "Features:\n"
            "- Import CSV/Excel/JSON data\n"
            "- Statistical analysis (ANOVA, post-hoc)\n"
            "- Publication-quality plots\n"
            "- Excel/GraphPad export\n\n"
            "Author: Jan Hawlicki\n"
            "Email: jan.hawlicki@uni-jena.de"
        )

    def run(self):
        """Start the application."""
        self.root.mainloop()

    def cleanup(self):
        """Clean up resources."""
        if self.database:
            self.database.disconnect()


def main():
    """Main entry point for GUI application."""
    root = tk.Tk()
    app = NeuromorphoAnalyzerApp(root)

    def on_close():
        app.cleanup()
        root.destroy()

    root.protocol("WM_DELETE_WINDOW", on_close)
    app.run()


if __name__ == "__main__":
    main()
