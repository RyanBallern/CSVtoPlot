"""Test suite for export functionality."""

import unittest
import tempfile
from pathlib import Path
import pandas as pd
import os
from openpyxl import load_workbook
import xml.etree.ElementTree as ET

from src.neuromorpho_analyzer.core.exporters import (
    ExportConfig,
    ExportParameterSelector,
    StatisticsTableExporter,
    ExcelExporter,
    GraphPadExporter
)
from src.neuromorpho_analyzer.core.processors.statistics import StatisticsEngine
from src.neuromorpho_analyzer.core.database.base import DatabaseBase


class TestExportConfig(unittest.TestCase):
    """Test ExportConfig class."""

    def test_default_config(self):
        """Test default configuration values."""
        config = ExportConfig()

        self.assertTrue(config.export_excel)
        self.assertFalse(config.export_graphpad)
        self.assertFalse(config.export_csv)
        self.assertTrue(config.export_statistics_tables)
        self.assertTrue(config.export_plots)
        self.assertEqual(config.plot_dpi, 800)
        self.assertEqual(config.plot_formats, ['png', 'tif'])

    def test_custom_config(self):
        """Test custom configuration."""
        config = ExportConfig(
            export_graphpad=True,
            plot_dpi=300,
            plot_formats=['png']
        )

        self.assertTrue(config.export_graphpad)
        self.assertEqual(config.plot_dpi, 300)
        self.assertEqual(config.plot_formats, ['png'])

    def test_plot_type_selection(self):
        """Test plot type selection."""
        config = ExportConfig()

        self.assertTrue(config.should_export_plot_type('boxplot_total'))
        self.assertTrue(config.should_export_plot_type('barplot_total'))
        self.assertFalse(config.should_export_plot_type('unknown_type'))

    def test_condition_selection(self):
        """Test condition selection."""
        # Test all conditions (default)
        config = ExportConfig()
        self.assertTrue(config.should_include_condition('Control'))
        self.assertTrue(config.should_include_condition('Treatment'))

        # Test specific conditions
        config = ExportConfig(selected_conditions=['Control'])
        self.assertTrue(config.should_include_condition('Control'))
        self.assertFalse(config.should_include_condition('Treatment'))

    def test_get_active_plot_types(self):
        """Test getting active plot types."""
        config = ExportConfig()
        plot_types = config.get_active_plot_types()

        self.assertIsInstance(plot_types, list)
        self.assertGreater(len(plot_types), 0)
        self.assertIn('boxplot_total', plot_types)


class MockDatabase(DatabaseBase):
    """Mock database for testing."""

    def __init__(self):
        self.measurements_df = pd.DataFrame({
            'assay_id': [1, 1, 1, 1, 2, 2, 2, 2],
            'condition': ['Control', 'Control', 'Treatment', 'Treatment',
                         'Control', 'Control', 'Treatment', 'Treatment'],
            'parameter_name': ['Length', 'Length', 'Length', 'Length',
                             'Width', 'Width', 'Width', 'Width'],
            'value': [10.5, 11.2, 15.3, 14.8, 5.2, 5.5, 7.1, 6.9]
        })

    def connect(self):
        """Connect to database."""
        pass

    def disconnect(self):
        """Disconnect from database."""
        pass

    def create_tables(self):
        """Create tables."""
        pass

    def insert_assay(self, name, description=None):
        """Insert assay."""
        return 1

    def get_assay(self, assay_id):
        """Get assay."""
        return {'id': assay_id, 'name': 'Test Assay'}

    def get_assay_by_name(self, name):
        """Get assay by name."""
        return {'id': 1, 'name': name}

    def list_assays(self):
        """List assays."""
        return [{'id': 1, 'name': 'Test Assay'}]

    def insert_measurements(self, assay_id, measurements, source_file=None, condition=None, check_duplicates=True):
        """Insert measurements."""
        return len(measurements)

    def delete_assay(self, assay_id):
        """Delete assay."""
        pass

    def get_measurement_count(self, assay_id):
        """Get measurement count."""
        df = self.measurements_df[self.measurements_df['assay_id'] == assay_id]
        return len(df)

    def get_conditions(self, assay_id):
        """Get conditions."""
        df = self.measurements_df[self.measurements_df['assay_id'] == assay_id]
        return df['condition'].unique().tolist()

    def get_parameters(self, assay_id):
        """Get parameters for an assay."""
        df = self.measurements_df[self.measurements_df['assay_id'] == assay_id]
        return df['parameter_name'].unique().tolist()

    def get_measurements(self, assay_id, condition=None, parameters=None):
        """Get measurements for an assay."""
        df = self.measurements_df[self.measurements_df['assay_id'] == assay_id]

        if condition:
            df = df[df['condition'] == condition]

        if parameters:
            df = df[df['parameter_name'].isin(parameters)]

        return df


class TestExportParameterSelector(unittest.TestCase):
    """Test ExportParameterSelector class."""

    def setUp(self):
        """Set up test fixtures."""
        self.db = MockDatabase()
        self.selector = ExportParameterSelector(self.db)

    def test_get_available_parameters(self):
        """Test getting available parameters."""
        params = self.selector.get_available_parameters([1, 2])

        self.assertEqual(len(params), 2)
        self.assertIn('Length', params)
        self.assertIn('Width', params)

    def test_select_parameters(self):
        """Test selecting parameters."""
        self.selector.select_parameters(['Length'])

        selected = self.selector.get_selected()
        self.assertEqual(len(selected), 1)
        self.assertIn('Length', selected)

    def test_select_all(self):
        """Test selecting all parameters."""
        self.selector.select_all([1, 2])

        selected = self.selector.get_selected()
        self.assertEqual(len(selected), 2)
        self.assertIn('Length', selected)
        self.assertIn('Width', selected)

    def test_is_selected(self):
        """Test checking if parameter is selected."""
        self.selector.select_parameters(['Length'])

        self.assertTrue(self.selector.is_selected('Length'))
        self.assertFalse(self.selector.is_selected('Width'))

    def test_toggle_parameter(self):
        """Test toggling parameter selection."""
        self.selector.select_parameters(['Length'])

        # Toggle off
        self.selector.toggle_parameter('Length')
        self.assertFalse(self.selector.is_selected('Length'))

        # Toggle on
        self.selector.toggle_parameter('Length')
        self.assertTrue(self.selector.is_selected('Length'))

    def test_clear_selection(self):
        """Test clearing selection."""
        self.selector.select_parameters(['Length', 'Width'])
        self.assertEqual(len(self.selector.get_selected()), 2)

        self.selector.clear_selection()
        self.assertEqual(len(self.selector.get_selected()), 0)


class TestStatisticsTableExporter(unittest.TestCase):
    """Test StatisticsTableExporter class."""

    def setUp(self):
        """Set up test fixtures."""
        self.stats_engine = StatisticsEngine(alpha=0.05)
        self.exporter = StatisticsTableExporter(self.stats_engine)

        # Create test data
        self.test_data = {
            'Control': pd.Series([10.5, 11.2, 10.8, 11.5]),
            'Treatment': pd.Series([15.3, 14.8, 15.1, 14.5])
        }

    def test_create_statistics_tables(self):
        """Test creating statistics tables."""
        tables = self.exporter.create_statistics_tables(
            self.test_data, 'Test Parameter'
        )

        self.assertIn('summary', tables)
        self.assertIn('anova', tables)
        self.assertIn('pairwise', tables)

        # Check summary table
        summary = tables['summary']
        self.assertEqual(len(summary), 2)  # 2 conditions
        self.assertIn('Parameter', summary.columns)
        self.assertIn('Condition', summary.columns)
        self.assertIn('Mean', summary.columns)

    def test_export_to_excel(self):
        """Test exporting to Excel."""
        tables = self.exporter.create_statistics_tables(
            self.test_data, 'Test Parameter'
        )

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / 'test_stats.xlsx'
            self.exporter.export_to_excel(tables, output_path)

            self.assertTrue(output_path.exists())

            # Load and verify
            wb = load_workbook(output_path)
            self.assertIn('Summary', wb.sheetnames)
            self.assertIn('Anova', wb.sheetnames)
            self.assertIn('Pairwise', wb.sheetnames)


class TestExcelExporter(unittest.TestCase):
    """Test ExcelExporter class."""

    def setUp(self):
        """Set up test fixtures."""
        self.db = MockDatabase()
        self.stats_engine = StatisticsEngine(alpha=0.05)
        self.selector = ExportParameterSelector(self.db)
        self.selector.select_all([1, 2])
        self.exporter = ExcelExporter(self.selector, self.stats_engine)

    def test_export(self):
        """Test exporting to Excel."""
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = self.exporter.export(
                [1, 2], Path(tmpdir), self.db
            )

            self.assertTrue(output_path.exists())
            self.assertTrue(output_path.name.startswith('analysis_'))
            self.assertTrue(output_path.name.endswith('.xlsx'))

            # Load and verify
            wb = load_workbook(output_path)
            self.assertIn('Raw Data', wb.sheetnames)
            self.assertIn('Summary Statistics', wb.sheetnames)
            self.assertIn('Frequency Distributions', wb.sheetnames)

    def test_raw_data_sheet(self):
        """Test raw data sheet creation."""
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = self.exporter.export(
                [1, 2], Path(tmpdir), self.db
            )

            wb = load_workbook(output_path)
            ws = wb['Raw Data']

            # Check headers
            headers = [cell.value for cell in ws[1]]
            self.assertIn('condition', headers)
            self.assertIn('parameter_name', headers)
            self.assertIn('value', headers)

    def test_summary_stats_sheet(self):
        """Test summary statistics sheet creation."""
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = self.exporter.export(
                [1, 2], Path(tmpdir), self.db
            )

            wb = load_workbook(output_path)
            ws = wb['Summary Statistics']

            # Check headers
            headers = [cell.value for cell in ws[1]]
            self.assertIn('Parameter', headers)
            self.assertIn('Condition', headers)
            self.assertIn('Mean', headers)
            self.assertIn('SEM', headers)


class TestGraphPadExporter(unittest.TestCase):
    """Test GraphPadExporter class."""

    def setUp(self):
        """Set up test fixtures."""
        self.db = MockDatabase()
        self.selector = ExportParameterSelector(self.db)
        self.selector.select_all([1, 2])
        self.exporter = GraphPadExporter(self.selector)

    def test_export(self):
        """Test exporting to GraphPad format."""
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = self.exporter.export(
                [1, 2], Path(tmpdir), self.db
            )

            self.assertTrue(output_path.exists())
            self.assertTrue(output_path.name.startswith('graphpad_'))
            self.assertTrue(output_path.name.endswith('.pzfx'))

    def test_xml_structure(self):
        """Test XML structure of exported file."""
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = self.exporter.export(
                [1, 2], Path(tmpdir), self.db
            )

            # Parse XML
            tree = ET.parse(output_path)
            root = tree.getroot()

            # Check root element (handle namespace)
            self.assertTrue(root.tag.endswith('GraphPadPrismFile'))

            # Check tables exist (handle namespace)
            tables = root.findall('.//{http://graphpad.com/prism/Prism.htm}Table')
            self.assertGreater(len(tables), 0)

    def test_parameter_tables(self):
        """Test parameter table creation."""
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = self.exporter.export(
                [1, 2], Path(tmpdir), self.db
            )

            # Parse XML
            tree = ET.parse(output_path)
            root = tree.getroot()

            # Find tables
            tables = root.findall('.//Table')

            # Check each table has title and columns
            for table in tables:
                title = table.find('Title')
                self.assertIsNotNone(title)

                columns = table.findall('YColumn')
                self.assertGreater(len(columns), 0)


def run_tests():
    """Run all export tests."""
    # Create test suite
    loader = unittest.TestLoader()
    suite = unittest.TestSuite()

    # Add all test classes
    suite.addTests(loader.loadTestsFromTestCase(TestExportConfig))
    suite.addTests(loader.loadTestsFromTestCase(TestExportParameterSelector))
    suite.addTests(loader.loadTestsFromTestCase(TestStatisticsTableExporter))
    suite.addTests(loader.loadTestsFromTestCase(TestExcelExporter))
    suite.addTests(loader.loadTestsFromTestCase(TestGraphPadExporter))

    # Run tests
    runner = unittest.TextTestRunner(verbosity=2)
    result = runner.run(suite)

    return result.wasSuccessful()


if __name__ == '__main__':
    success = run_tests()
    exit(0 if success else 1)
