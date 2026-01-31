"""Simple diagnostic script to test export functionality step by step."""

import sys
from pathlib import Path

print("="*60)
print("Export Functionality Diagnostic Test")
print("="*60)

# Test 1: Import modules
print("\n[1/6] Testing imports...")
try:
    from src.neuromorpho_analyzer.core.exporters import (
        ExportConfig,
        ExportParameterSelector,
        StatisticsTableExporter,
        ExcelExporter,
        GraphPadExporter
    )
    print("✓ All export modules imported successfully")
except ImportError as e:
    print(f"✗ Import error: {e}")
    sys.exit(1)

# Test 2: Import dependencies
print("\n[2/6] Testing dependencies...")
try:
    from src.neuromorpho_analyzer.core.processors.statistics import StatisticsEngine
    from src.neuromorpho_analyzer.core.database.sqlite import SQLiteDatabase
    import pandas as pd
    import numpy as np
    print("✓ All dependencies imported successfully")
except ImportError as e:
    print(f"✗ Dependency error: {e}")
    sys.exit(1)

# Test 3: Create ExportConfig
print("\n[3/6] Testing ExportConfig...")
try:
    config = ExportConfig()
    print(f"✓ ExportConfig created")
    print(f"  - Export Excel: {config.export_excel}")
    print(f"  - Plot DPI: {config.plot_dpi}")
    print(f"  - Active plot types: {len(config.get_active_plot_types())}")
except Exception as e:
    print(f"✗ ExportConfig error: {e}")
    sys.exit(1)

# Test 4: Create test database
print("\n[4/6] Testing database creation...")
try:
    import tempfile
    db_path = Path(tempfile.gettempdir()) / 'diagnostic_test.db'
    if db_path.exists():
        db_path.unlink()

    db = SQLiteDatabase(str(db_path))
    db.connect()
    db.create_tables()

    # Insert test data
    assay_id = db.insert_assay('Test', 'Diagnostic test')

    test_data = pd.DataFrame({
        'parameter_name': ['Length'] * 10,
        'value': np.random.normal(100, 15, 10)
    })

    db.insert_measurements(
        assay_id=assay_id,
        measurements=test_data,
        condition='Control',
        check_duplicates=False
    )

    print(f"✓ Database created at: {db_path}")
    print(f"  - Assay ID: {assay_id}")
    print(f"  - Measurements inserted: {len(test_data)}")

    # Test retrieval
    retrieved = db.get_measurements(assay_id)
    print(f"  - Measurements retrieved: {len(retrieved)}")

except Exception as e:
    print(f"✗ Database error: {e}")
    import traceback
    traceback.print_exc()
    sys.exit(1)

# Test 5: Test parameter selector
print("\n[5/6] Testing ExportParameterSelector...")
try:
    selector = ExportParameterSelector(db)
    available = selector.get_available_parameters([assay_id])
    print(f"✓ Parameter selector created")
    print(f"  - Available parameters: {available}")

    selector.select_all([assay_id])
    selected = selector.get_selected()
    print(f"  - Selected parameters: {selected}")

except Exception as e:
    print(f"✗ Parameter selector error: {e}")
    import traceback
    traceback.print_exc()
    sys.exit(1)

# Test 6: Test Excel export
print("\n[6/6] Testing Excel export...")
try:
    output_dir = Path(tempfile.gettempdir())

    stats_engine = StatisticsEngine(alpha=0.05)
    exporter = ExcelExporter(selector, stats_engine)

    output_file = exporter.export([assay_id], output_dir, db)

    print(f"✓ Excel export successful")
    print(f"  - Output file: {output_file}")
    print(f"  - File size: {output_file.stat().st_size:,} bytes")
    print(f"  - File exists: {output_file.exists()}")

    # Verify file content
    from openpyxl import load_workbook
    wb = load_workbook(output_file)
    print(f"  - Worksheets: {wb.sheetnames}")

except Exception as e:
    print(f"✗ Excel export error: {e}")
    import traceback
    traceback.print_exc()
    sys.exit(1)

finally:
    db.disconnect()
    if db_path.exists():
        db_path.unlink()
    if output_file.exists():
        output_file.unlink()

print("\n" + "="*60)
print("✓ All diagnostic tests passed!")
print("="*60)
print("\nThe export functionality is working correctly.")
print("\nTo run the full test suite:")
print("  python test_export.py")
print("\nTo run the demo with example exports:")
print("  python demo_export.py")
