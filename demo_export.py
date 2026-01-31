"""Demo script to test export functionality."""

import pandas as pd
import numpy as np
from pathlib import Path
import tempfile

from src.neuromorpho_analyzer.core.exporters import (
    ExportConfig,
    ExportParameterSelector,
    StatisticsTableExporter,
    ExcelExporter,
    GraphPadExporter
)
from src.neuromorpho_analyzer.core.processors.statistics import StatisticsEngine
from src.neuromorpho_analyzer.core.database.sqlite import SQLiteDatabase


def create_test_data():
    """Create test database with sample measurements."""
    # Create temporary database
    db_path = Path(tempfile.gettempdir()) / 'test_export.db'
    if db_path.exists():
        db_path.unlink()

    db = SQLiteDatabase(str(db_path))
    db.connect()
    db.create_tables()

    # Create test assay
    assay_id = db.insert_assay('Test Analysis', 'Demo export data')

    # Generate sample data for multiple conditions and parameters
    np.random.seed(42)

    conditions = ['Control', 'Treatment A', 'Treatment B']
    parameters = ['Neurite Length', 'Branch Points', 'Cell Body Area']

    for condition in conditions:
        for param in parameters:
            # Generate realistic data with different means per condition
            if condition == 'Control':
                base_mean = 100
            elif condition == 'Treatment A':
                base_mean = 130
            else:  # Treatment B
                base_mean = 85

            # Adjust mean by parameter
            if param == 'Neurite Length':
                mean = base_mean
                std = 15
            elif param == 'Branch Points':
                mean = base_mean * 0.5
                std = 8
            else:  # Cell Body Area
                mean = base_mean * 1.2
                std = 20

            # Generate 20 measurements
            values = np.random.normal(mean, std, 20)

            # Create DataFrame in expected format
            measurements = pd.DataFrame({
                'parameter_name': [param] * len(values),
                'value': values
            })

            # Insert measurements
            db.insert_measurements(
                assay_id=assay_id,
                measurements=measurements,
                source_file=f'{condition}.xlsx',
                condition=condition,
                check_duplicates=False
            )

    print(f"✓ Created test database: {db_path}")
    print(f"  - Assay ID: {assay_id}")
    print(f"  - Conditions: {len(conditions)}")
    print(f"  - Parameters: {len(parameters)}")
    print(f"  - Total measurements: {len(conditions) * len(parameters) * 20}")

    return db, assay_id


def test_statistics_table_export(db, assay_id):
    """Test StatisticsTableExporter."""
    print("\n" + "="*60)
    print("Testing StatisticsTableExporter")
    print("="*60)

    # Create output directory relative to current working directory
    output_dir = Path.cwd() / 'demo_output'
    output_dir.mkdir(parents=True, exist_ok=True)

    # Get data for one parameter
    param = 'Neurite Length'
    df = db.get_measurements(assay_id)
    param_data = df[df['parameter_name'] == param]

    # Organize by condition
    data_dict = {}
    for condition in param_data['condition'].unique():
        cond_data = param_data[param_data['condition'] == condition]['value']
        data_dict[condition] = pd.Series(cond_data.values)

    # Create exporter
    stats_engine = StatisticsEngine(alpha=0.05)
    exporter = StatisticsTableExporter(stats_engine)

    # Create tables
    print(f"\nCreating statistics tables for '{param}'...")
    tables = exporter.create_statistics_tables(data_dict, param)

    print(f"\n✓ Created {len(tables)} tables:")
    for table_name, table_df in tables.items():
        print(f"  - {table_name}: {len(table_df)} rows")
        print(f"\n    {table_name.upper()} Preview:")
        print(f"    {'-'*56}")
        print(table_df.to_string(index=False, max_rows=5).replace('\n', '\n    '))

    # Export to Excel
    output_file = output_dir / f'statistics_tables_{param.replace(" ", "_")}.xlsx'
    exporter.export_to_excel(tables, output_file)
    print(f"\n✓ Exported to: {output_file}")
    print(f"  File size: {output_file.stat().st_size:,} bytes")


def test_excel_export(db, assay_id):
    """Test ExcelExporter."""
    print("\n" + "="*60)
    print("Testing ExcelExporter")
    print("="*60)

    # Create output directory relative to current working directory
    output_dir = Path.cwd() / 'demo_output'
    output_dir.mkdir(parents=True, exist_ok=True)

    # Create parameter selector and select all parameters
    param_selector = ExportParameterSelector(db)
    param_selector.select_all([assay_id])

    print(f"\nSelected parameters:")
    for param in param_selector.get_selected():
        print(f"  - {param}")

    # Create exporter
    stats_engine = StatisticsEngine(alpha=0.05)
    exporter = ExcelExporter(param_selector, stats_engine)

    # Export
    print(f"\nExporting comprehensive Excel file...")
    output_file = exporter.export([assay_id], output_dir, db)

    print(f"\n✓ Exported to: {output_file}")
    print(f"  File size: {output_file.stat().st_size:,} bytes")

    # Show what sheets were created
    from openpyxl import load_workbook
    wb = load_workbook(output_file)
    print(f"\n  Worksheets created:")
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = ws.max_row
        cols = ws.max_column
        print(f"    - {sheet_name}: {rows} rows × {cols} columns")


def test_graphpad_export(db, assay_id):
    """Test GraphPadExporter."""
    print("\n" + "="*60)
    print("Testing GraphPadExporter")
    print("="*60)

    # Create output directory relative to current working directory
    output_dir = Path.cwd() / 'demo_output'
    output_dir.mkdir(parents=True, exist_ok=True)

    # Create parameter selector and select all parameters
    param_selector = ExportParameterSelector(db)
    param_selector.select_all([assay_id])

    print(f"\nSelected parameters:")
    for param in param_selector.get_selected():
        print(f"  - {param}")

    # Create exporter
    exporter = GraphPadExporter(param_selector)

    # Export
    print(f"\nExporting GraphPad Prism file...")
    output_file = exporter.export([assay_id], output_dir, db)

    print(f"\n✓ Exported to: {output_file}")
    print(f"  File size: {output_file.stat().st_size:,} bytes")

    # Show XML structure
    import xml.etree.ElementTree as ET
    tree = ET.parse(output_file)
    root = tree.getroot()

    tables = root.findall('.//{http://graphpad.com/prism/Prism.htm}Table')
    print(f"\n  Tables created: {len(tables)}")
    for table in tables:
        title_elem = table.find('{http://graphpad.com/prism/Prism.htm}Title')
        title = title_elem.text if title_elem is not None else 'Unknown'
        columns = table.findall('.//{http://graphpad.com/prism/Prism.htm}YColumn')
        print(f"    - {title}: {len(columns)} conditions")


def test_export_config():
    """Test ExportConfig."""
    print("\n" + "="*60)
    print("Testing ExportConfig")
    print("="*60)

    # Create default config
    config = ExportConfig()

    print("\nDefault configuration:")
    print(f"  Export Excel: {config.export_excel}")
    print(f"  Export GraphPad: {config.export_graphpad}")
    print(f"  Export CSV: {config.export_csv}")
    print(f"  Export Statistics Tables: {config.export_statistics_tables}")
    print(f"  Export Plots: {config.export_plots}")
    print(f"  Plot DPI: {config.plot_dpi}")
    print(f"  Plot Formats: {config.plot_formats}")

    print(f"\n  Plot Types:")
    for plot_type in config.get_active_plot_types():
        print(f"    - {plot_type}")

    # Create custom config
    print("\nCustom configuration:")
    custom_config = ExportConfig(
        export_graphpad=True,
        plot_dpi=300,
        plot_formats=['png'],
        selected_conditions=['Control', 'Treatment A']
    )

    print(f"  Export GraphPad: {custom_config.export_graphpad}")
    print(f"  Plot DPI: {custom_config.plot_dpi}")
    print(f"  Plot Formats: {custom_config.plot_formats}")

    print(f"\n  Condition filtering:")
    for condition in ['Control', 'Treatment A', 'Treatment B']:
        included = custom_config.should_include_condition(condition)
        print(f"    - {condition}: {'✓ included' if included else '✗ excluded'}")


def main():
    """Run all export tests."""
    print("\n" + "="*60)
    print("CSVtoPlot Analyzer - Export Functionality Demo")
    print("="*60)

    # Create test data
    db, assay_id = create_test_data()

    try:
        # Test export config
        test_export_config()

        # Test statistics table export
        test_statistics_table_export(db, assay_id)

        # Test Excel export
        test_excel_export(db, assay_id)

        # Test GraphPad export
        test_graphpad_export(db, assay_id)

        print("\n" + "="*60)
        print("All export tests completed successfully!")
        print("="*60)

        output_dir = Path.cwd() / 'demo_output'
        print(f"\nOutput files saved to: {output_dir.absolute()}")
        print("\nGenerated files:")
        for file in sorted(output_dir.glob('*')):
            if file.is_file() and file.stat().st_mtime > (file.stat().st_mtime - 60):
                print(f"  - {file.name} ({file.stat().st_size:,} bytes)")

    finally:
        db.disconnect()


if __name__ == '__main__':
    main()
