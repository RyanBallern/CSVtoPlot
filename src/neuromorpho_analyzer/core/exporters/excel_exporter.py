"""Export comprehensive analysis results to Excel matching the expected format."""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple
import pandas as pd
import numpy as np

from .parameter_selector import ExportParameterSelector
from ..processors.statistics import StatisticsEngine
from ..database.base import DatabaseBase


# Style constants matching the example file
HEADER_FILL = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
HEADER_FONT = Font(bold=True, size=11, color='FFFFFF')
SECTION_TITLE_FONT_LARGE = Font(bold=True, size=14)
SECTION_TITLE_FONT = Font(bold=True, size=12)
LABEL_FONT = Font(bold=True, size=11)
DEFAULT_FONT = Font(size=11)
CENTER_ALIGN = Alignment(horizontal='center', vertical='center')
LEFT_ALIGN = Alignment(horizontal='left', vertical='center')


class ExcelExporter:
    """Exports comprehensive analysis results to Excel in the standard format."""

    def __init__(self, parameter_selector: ExportParameterSelector,
                 stats_engine: StatisticsEngine):
        """Initialize Excel exporter.

        Args:
            parameter_selector: Parameter selector for export
            stats_engine: Statistics engine for calculations
        """
        self.param_selector = parameter_selector
        self.stats = stats_engine

    def export(self, assay_ids: List[int], output_dir: Path,
               database: DatabaseBase,
               dataset_split: Optional[Dict[str, str]] = None,
               bin_size: float = 10.0,
               bin_count: int = 250) -> Path:
        """Export comprehensive Excel file with all parameters and statistics.

        Args:
            assay_ids: List of assay IDs to export
            output_dir: Output directory
            database: Database interface
            dataset_split: Optional dict mapping markers (L/T) to names (Liposome/Tubule)
            bin_size: Size of bins for frequency distribution
            bin_count: Number of bins for frequency distribution

        Returns:
            Path to created Excel file
        """
        # Get data from all assays
        parameters = self.param_selector.get_selected()
        dfs = []
        for assay_id in assay_ids:
            assay_df = database.get_measurements(assay_id, parameters=parameters)
            if not assay_df.empty:
                assay_df['assay_id'] = assay_id
                dfs.append(assay_df)

        # Combine all data
        if dfs:
            df = pd.concat(dfs, ignore_index=True)
        else:
            df = pd.DataFrame()

        # Default dataset split if not provided
        if dataset_split is None:
            dataset_split = {'L': 'Liposome', 'T': 'Tubule'}

        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Generate filename
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        assay_indices = '_'.join(str(aid) for aid in sorted(assay_ids))
        filename = f'analysis_{timestamp}_assays{assay_indices}.xlsx'

        # Get conditions
        conditions = sorted(df['condition'].unique()) if not df.empty else []

        # Create sheets for each dataset type
        for marker, dataset_name in dataset_split.items():
            # Filter data by marker if available (check source_file for marker)
            dataset_df = self._filter_by_marker(df, marker)
            if dataset_df.empty:
                # If no marker-based data, use all data for first dataset
                if marker == list(dataset_split.keys())[0]:
                    dataset_df = df
                else:
                    continue

            # Create data sheet (raw values with stats header)
            self._create_data_sheet(wb, dataset_df, dataset_name, conditions)

            # Create relative change sheet
            self._create_relative_change_sheet(wb, dataset_df, dataset_name, conditions)

            # Create frequency sheet
            self._create_frequency_sheet(wb, dataset_df, dataset_name, conditions,
                                         bin_size=bin_size, bin_count=bin_count)

            # Create statistics sheet
            self._create_statistics_sheet(wb, dataset_df, dataset_name, conditions)

        # Save
        output_path = output_dir / filename
        wb.save(output_path)

        return output_path

    def _filter_by_marker(self, df: pd.DataFrame, marker: str) -> pd.DataFrame:
        """Filter DataFrame by dataset marker in source file name.

        Args:
            df: Input DataFrame
            marker: Marker character (L or T)

        Returns:
            Filtered DataFrame
        """
        if df.empty or 'source_file' not in df.columns:
            return pd.DataFrame()

        # Check if source file ends with marker before extension
        mask = df['source_file'].str.contains(f'{marker}\\.[^.]+$', regex=True, na=False)
        return df[mask].copy()

    def _create_data_sheet(self, wb: Workbook, df: pd.DataFrame,
                           sheet_name: str, conditions: List[str]) -> None:
        """Create data worksheet with conditions as columns.

        Format:
        Row 1: "Condition:" | cond1 | cond2 | ...
        Row 2: "Average:"   | avg1  | avg2  | ...
        Row 3: "SEM:"       | sem1  | sem2  | ...
        Row 4: "Count:"     | cnt1  | cnt2  | ...
        Row 5+: "Values:"   | val1  | val2  | ... (one row per value)
        """
        ws = wb.create_sheet(sheet_name)

        if df.empty:
            self._write_empty_data_sheet(ws, conditions)
            return

        # Get values grouped by condition
        condition_values = {}
        for cond in conditions:
            cond_data = df[df['condition'] == cond]['value'].dropna().tolist()
            condition_values[cond] = cond_data

        # Row 1: Condition headers
        ws.cell(row=1, column=1, value='Condition:').font = DEFAULT_FONT
        for col_idx, cond in enumerate(conditions, 2):
            ws.cell(row=1, column=col_idx, value=cond).font = DEFAULT_FONT

        # Row 2: Average
        ws.cell(row=2, column=1, value='Average:').font = DEFAULT_FONT
        for col_idx, cond in enumerate(conditions, 2):
            values = condition_values.get(cond, [])
            if values:
                avg = np.mean(values)
                ws.cell(row=2, column=col_idx, value=avg)

        # Row 3: SEM
        ws.cell(row=3, column=1, value='SEM:').font = DEFAULT_FONT
        for col_idx, cond in enumerate(conditions, 2):
            values = condition_values.get(cond, [])
            if len(values) > 1:
                sem = np.std(values, ddof=1) / np.sqrt(len(values))
                ws.cell(row=3, column=col_idx, value=sem)

        # Row 4: Count
        ws.cell(row=4, column=1, value='Count:').font = DEFAULT_FONT
        for col_idx, cond in enumerate(conditions, 2):
            values = condition_values.get(cond, [])
            ws.cell(row=4, column=col_idx, value=len(values))

        # Row 5+: Values
        max_values = max(len(v) for v in condition_values.values()) if condition_values else 0
        for row_idx in range(5, 5 + max_values):
            value_idx = row_idx - 5
            if value_idx == 0:
                ws.cell(row=row_idx, column=1, value='Values:')

            for col_idx, cond in enumerate(conditions, 2):
                values = condition_values.get(cond, [])
                if value_idx < len(values):
                    ws.cell(row=row_idx, column=col_idx, value=values[value_idx])

    def _write_empty_data_sheet(self, ws, conditions: List[str]) -> None:
        """Write empty data sheet structure."""
        ws.cell(row=1, column=1, value='Condition:').font = DEFAULT_FONT
        for col_idx, cond in enumerate(conditions, 2):
            ws.cell(row=1, column=col_idx, value=cond).font = DEFAULT_FONT

        ws.cell(row=2, column=1, value='Average:').font = DEFAULT_FONT
        ws.cell(row=3, column=1, value='SEM:').font = DEFAULT_FONT
        ws.cell(row=4, column=1, value='Count:').font = DEFAULT_FONT
        ws.cell(row=5, column=1, value='Values:')

    def _create_relative_change_sheet(self, wb: Workbook, df: pd.DataFrame,
                                       base_name: str, conditions: List[str]) -> None:
        """Create relative change worksheet."""
        sheet_name = f'{base_name}_RelativeChange'
        ws = wb.create_sheet(sheet_name)

        # Same structure as data sheet but for relative change values
        ws.cell(row=1, column=1, value='Condition:').font = DEFAULT_FONT
        for col_idx, cond in enumerate(conditions, 2):
            ws.cell(row=1, column=col_idx, value=cond).font = DEFAULT_FONT

        ws.cell(row=2, column=1, value='Average:').font = DEFAULT_FONT
        ws.cell(row=3, column=1, value='SEM:').font = DEFAULT_FONT
        ws.cell(row=4, column=1, value='Count:').font = DEFAULT_FONT
        ws.cell(row=5, column=1, value='Values:')

        # Relative change would be calculated against control - placeholder for now

    def _create_frequency_sheet(self, wb: Workbook, df: pd.DataFrame,
                                 base_name: str, conditions: List[str],
                                 bin_size: float = 10.0,
                                 bin_count: int = 250) -> None:
        """Create frequency distribution worksheet.

        Format:
        Header: Bin Center | cond1_Frequency | cond1_Percentage | cond2_Frequency | ...
        Data rows with frequency counts and percentages
        """
        sheet_name = f'{base_name}_Frequency'
        ws = wb.create_sheet(sheet_name)

        # Create headers
        headers = ['Bin Center']
        for cond in conditions:
            headers.extend([f'{cond}_Frequency', f'{cond}_Percentage'])

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL

        if df.empty:
            return

        # Calculate bin centers (starting from bin_size/2 + 10)
        bin_start = 10 + bin_size / 2
        bin_centers = [bin_start + i * bin_size for i in range(bin_count)]

        # Get frequency data for each condition
        condition_freqs = {}
        for cond in conditions:
            cond_values = df[df['condition'] == cond]['value'].dropna().values
            if len(cond_values) > 0:
                # Calculate histogram
                bins = [10 + i * bin_size for i in range(bin_count + 1)]
                hist, _ = np.histogram(cond_values, bins=bins)
                total = len(cond_values)
                percentages = (hist / total * 100) if total > 0 else hist
                condition_freqs[cond] = (hist, percentages)
            else:
                condition_freqs[cond] = (np.zeros(bin_count), np.zeros(bin_count))

        # Write data rows
        for row_idx, bin_center in enumerate(bin_centers, 2):
            ws.cell(row=row_idx, column=1, value=int(bin_center))

            col_idx = 2
            for cond in conditions:
                hist, pct = condition_freqs.get(cond, (np.zeros(bin_count), np.zeros(bin_count)))
                bin_idx = row_idx - 2
                if bin_idx < len(hist):
                    freq = int(hist[bin_idx])
                    percentage = pct[bin_idx]
                    # Only write non-zero values
                    if freq > 0:
                        ws.cell(row=row_idx, column=col_idx, value=freq)
                        ws.cell(row=row_idx, column=col_idx + 1, value=round(percentage, 2))
                col_idx += 2

    def _create_statistics_sheet(self, wb: Workbook, df: pd.DataFrame,
                                  base_name: str, conditions: List[str]) -> None:
        """Create statistics worksheet with three sections.

        Format:
        Row 1: "STATISTICAL ANALYSIS" (bold, 14pt)
        Row 3-6: Overall test info
        Row 9: "DESCRIPTIVE STATISTICS" (bold, 12pt)
        Row 10: Header row (bold, blue fill)
        Row 11+: Data rows
        Row N: "PAIRWISE COMPARISONS" (bold, 12pt)
        Row N+1: Header row (bold, blue fill)
        Row N+2+: Pairwise comparison data
        """
        sheet_name = f'{base_name}_Statistics'
        ws = wb.create_sheet(sheet_name)

        if df.empty:
            self._write_empty_statistics_sheet(ws)
            return

        # Prepare data for statistics
        data_dict = {}
        for cond in conditions:
            cond_values = df[df['condition'] == cond]['value'].dropna()
            if len(cond_values) > 0:
                data_dict[cond] = cond_values

        if not data_dict:
            self._write_empty_statistics_sheet(ws)
            return

        # Convert to DataFrame for statistics
        stats_df = pd.DataFrame([
            {'condition': cond, 'value': val}
            for cond, series in data_dict.items()
            for val in series
        ])

        # Run statistical analysis
        try:
            result = self.stats.auto_compare(stats_df, 'value', 'condition')
        except Exception as e:
            self._write_empty_statistics_sheet(ws, error=str(e))
            return

        current_row = 1

        # Section 1: STATISTICAL ANALYSIS
        ws.cell(row=current_row, column=1, value='STATISTICAL ANALYSIS').font = SECTION_TITLE_FONT_LARGE
        ws.row_dimensions[current_row].height = 18.75
        current_row += 2

        # Overall test info
        main_test = result.get('main_test')
        if main_test:
            ws.cell(row=current_row, column=1, value='Overall Test:').font = LABEL_FONT
            test_name = main_test.test_name
            if result.get('is_parametric') is False:
                test_name += ' (non-parametric)'
            ws.cell(row=current_row, column=2, value=test_name)
            current_row += 1

            ws.cell(row=current_row, column=1, value='Test Statistic:')
            ws.cell(row=current_row, column=2, value=round(main_test.statistic, 4))
            current_row += 1

            ws.cell(row=current_row, column=1, value='P-value:')
            ws.cell(row=current_row, column=2, value=f'{main_test.p_value:.4e}')
            current_row += 1

            ws.cell(row=current_row, column=1, value='Significant (α=0.05):')
            ws.cell(row=current_row, column=2, value='Yes' if main_test.significant else 'No')
            current_row += 3

        # Section 2: DESCRIPTIVE STATISTICS
        ws.cell(row=current_row, column=1, value='DESCRIPTIVE STATISTICS').font = SECTION_TITLE_FONT
        ws.row_dimensions[current_row].height = 15.75
        current_row += 1

        # Header row
        desc_headers = ['Condition', 'N', 'Mean', 'Median', 'SD', 'SEM', 'Min', 'Max']
        for col_idx, header in enumerate(desc_headers, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
        current_row += 1

        # Descriptive stats data
        for cond, series in data_dict.items():
            n = len(series)
            mean = series.mean()
            median = series.median()
            std = series.std(ddof=1)
            sem = std / np.sqrt(n) if n > 0 else 0
            min_val = series.min()
            max_val = series.max()

            values = [cond, n, round(mean, 3), round(median, 3),
                      round(std, 3), round(sem, 3), round(min_val, 3), round(max_val, 3)]
            for col_idx, val in enumerate(values, 1):
                ws.cell(row=current_row, column=col_idx, value=val)
            current_row += 1

        current_row += 2

        # Section 3: PAIRWISE COMPARISONS
        ws.cell(row=current_row, column=1, value='PAIRWISE COMPARISONS').font = SECTION_TITLE_FONT
        ws.row_dimensions[current_row].height = 15.75
        current_row += 1

        # Header row
        pairwise_headers = ['Group 1', 'Group 2', 'Test', 'Statistic', 'P-value', 'Significant']
        for col_idx, header in enumerate(pairwise_headers, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
        current_row += 1

        # Pairwise comparison data
        post_hoc = result.get('post_hoc_tests', [])
        if post_hoc:
            for test in post_hoc:
                test_name = 'Mann-Whitney U' if not result.get('is_parametric', True) else 'Tukey HSD'
                significant = 'Yes' if test.p_value < 0.05 else 'No'

                values = [test.group1, test.group2, test_name,
                          round(test.statistic, 4) if hasattr(test, 'statistic') else '',
                          f'{test.p_value:.4e}', significant]
                for col_idx, val in enumerate(values, 1):
                    ws.cell(row=current_row, column=col_idx, value=val)
                current_row += 1

    def _write_empty_statistics_sheet(self, ws, error: str = None) -> None:
        """Write empty statistics sheet structure."""
        ws.cell(row=1, column=1, value='STATISTICAL ANALYSIS').font = SECTION_TITLE_FONT_LARGE
        ws.row_dimensions[1].height = 18.75

        if error:
            ws.cell(row=3, column=1, value='Error:')
            ws.cell(row=3, column=2, value=error)

        ws.cell(row=9, column=1, value='DESCRIPTIVE STATISTICS').font = SECTION_TITLE_FONT
        ws.row_dimensions[9].height = 15.75

        desc_headers = ['Condition', 'N', 'Mean', 'Median', 'SD', 'SEM', 'Min', 'Max']
        for col_idx, header in enumerate(desc_headers, 1):
            cell = ws.cell(row=10, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL

        ws.cell(row=19, column=1, value='PAIRWISE COMPARISONS').font = SECTION_TITLE_FONT
        ws.row_dimensions[19].height = 15.75

        pairwise_headers = ['Group 1', 'Group 2', 'Test', 'Statistic', 'P-value', 'Significant']
        for col_idx, header in enumerate(pairwise_headers, 1):
            cell = ws.cell(row=20, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL


class ExcelExporterSimple:
    """Simpler Excel exporter for single-dataset exports."""

    def __init__(self, stats_engine: StatisticsEngine = None):
        """Initialize simple Excel exporter.

        Args:
            stats_engine: Optional statistics engine for calculations
        """
        self.stats = stats_engine or StatisticsEngine()

    def export_data(self, data: Dict[str, List[float]], output_path: Path,
                    dataset_name: str = 'Data',
                    include_statistics: bool = True,
                    include_frequency: bool = True,
                    bin_size: float = 10.0,
                    bin_count: int = 250) -> Path:
        """Export data dictionary to Excel file.

        Args:
            data: Dictionary mapping condition names to lists of values
            output_path: Output file path
            dataset_name: Name for the dataset sheet
            include_statistics: Whether to include statistics sheet
            include_frequency: Whether to include frequency sheet
            bin_size: Size of bins for frequency distribution
            bin_count: Number of bins for frequency distribution

        Returns:
            Path to created Excel file
        """
        wb = Workbook()
        wb.remove(wb.active)

        conditions = list(data.keys())

        # Create data sheet
        self._create_data_sheet_from_dict(wb, data, dataset_name, conditions)

        # Create frequency sheet
        if include_frequency:
            self._create_frequency_sheet_from_dict(wb, data, dataset_name, conditions,
                                                    bin_size, bin_count)

        # Create statistics sheet
        if include_statistics:
            self._create_statistics_sheet_from_dict(wb, data, dataset_name, conditions)

        wb.save(output_path)
        return output_path

    def _create_data_sheet_from_dict(self, wb: Workbook, data: Dict[str, List[float]],
                                      sheet_name: str, conditions: List[str]) -> None:
        """Create data sheet from dictionary."""
        ws = wb.create_sheet(sheet_name)

        # Row 1: Condition headers
        ws.cell(row=1, column=1, value='Condition:').font = DEFAULT_FONT
        for col_idx, cond in enumerate(conditions, 2):
            ws.cell(row=1, column=col_idx, value=cond).font = DEFAULT_FONT

        # Row 2: Average
        ws.cell(row=2, column=1, value='Average:').font = DEFAULT_FONT
        for col_idx, cond in enumerate(conditions, 2):
            values = data.get(cond, [])
            if values:
                ws.cell(row=2, column=col_idx, value=np.mean(values))

        # Row 3: SEM
        ws.cell(row=3, column=1, value='SEM:').font = DEFAULT_FONT
        for col_idx, cond in enumerate(conditions, 2):
            values = data.get(cond, [])
            if len(values) > 1:
                ws.cell(row=3, column=col_idx, value=np.std(values, ddof=1) / np.sqrt(len(values)))

        # Row 4: Count
        ws.cell(row=4, column=1, value='Count:').font = DEFAULT_FONT
        for col_idx, cond in enumerate(conditions, 2):
            ws.cell(row=4, column=col_idx, value=len(data.get(cond, [])))

        # Row 5+: Values
        max_values = max(len(v) for v in data.values()) if data else 0
        for row_idx in range(5, 5 + max_values):
            value_idx = row_idx - 5
            if value_idx == 0:
                ws.cell(row=row_idx, column=1, value='Values:')

            for col_idx, cond in enumerate(conditions, 2):
                values = data.get(cond, [])
                if value_idx < len(values):
                    ws.cell(row=row_idx, column=col_idx, value=values[value_idx])

    def _create_frequency_sheet_from_dict(self, wb: Workbook, data: Dict[str, List[float]],
                                           base_name: str, conditions: List[str],
                                           bin_size: float, bin_count: int) -> None:
        """Create frequency sheet from dictionary."""
        sheet_name = f'{base_name}_Frequency'
        ws = wb.create_sheet(sheet_name)

        # Headers
        headers = ['Bin Center']
        for cond in conditions:
            headers.extend([f'{cond}_Frequency', f'{cond}_Percentage'])

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL

        # Bin centers
        bin_start = 10 + bin_size / 2
        bin_centers = [bin_start + i * bin_size for i in range(bin_count)]

        # Calculate frequencies
        condition_freqs = {}
        for cond in conditions:
            values = np.array(data.get(cond, []))
            if len(values) > 0:
                bins = [10 + i * bin_size for i in range(bin_count + 1)]
                hist, _ = np.histogram(values, bins=bins)
                total = len(values)
                pct = (hist / total * 100) if total > 0 else hist
                condition_freqs[cond] = (hist, pct)
            else:
                condition_freqs[cond] = (np.zeros(bin_count), np.zeros(bin_count))

        # Write data
        for row_idx, bin_center in enumerate(bin_centers, 2):
            ws.cell(row=row_idx, column=1, value=int(bin_center))

            col_idx = 2
            for cond in conditions:
                hist, pct = condition_freqs.get(cond, (np.zeros(bin_count), np.zeros(bin_count)))
                bin_idx = row_idx - 2
                if bin_idx < len(hist) and hist[bin_idx] > 0:
                    ws.cell(row=row_idx, column=col_idx, value=int(hist[bin_idx]))
                    ws.cell(row=row_idx, column=col_idx + 1, value=round(pct[bin_idx], 2))
                col_idx += 2

    def _create_statistics_sheet_from_dict(self, wb: Workbook, data: Dict[str, List[float]],
                                            base_name: str, conditions: List[str]) -> None:
        """Create statistics sheet from dictionary."""
        sheet_name = f'{base_name}_Statistics'
        ws = wb.create_sheet(sheet_name)

        # Convert to DataFrame for statistics
        stats_df = pd.DataFrame([
            {'condition': cond, 'value': val}
            for cond, values in data.items()
            for val in values
        ])

        current_row = 1

        # Section 1: STATISTICAL ANALYSIS
        ws.cell(row=current_row, column=1, value='STATISTICAL ANALYSIS').font = SECTION_TITLE_FONT_LARGE
        ws.row_dimensions[current_row].height = 18.75
        current_row += 2

        # Run statistics
        try:
            result = self.stats.auto_compare(stats_df, 'value', 'condition')
            main_test = result.get('main_test')

            if main_test:
                ws.cell(row=current_row, column=1, value='Overall Test:').font = LABEL_FONT
                test_name = main_test.test_name
                if result.get('is_parametric') is False:
                    test_name += ' (non-parametric)'
                ws.cell(row=current_row, column=2, value=test_name)
                current_row += 1

                ws.cell(row=current_row, column=1, value='Test Statistic:')
                ws.cell(row=current_row, column=2, value=round(main_test.statistic, 4))
                current_row += 1

                ws.cell(row=current_row, column=1, value='P-value:')
                ws.cell(row=current_row, column=2, value=f'{main_test.p_value:.4e}')
                current_row += 1

                ws.cell(row=current_row, column=1, value='Significant (α=0.05):')
                ws.cell(row=current_row, column=2, value='Yes' if main_test.significant else 'No')
        except Exception as e:
            ws.cell(row=current_row, column=1, value='Error:')
            ws.cell(row=current_row, column=2, value=str(e))

        current_row = 9

        # Section 2: DESCRIPTIVE STATISTICS
        ws.cell(row=current_row, column=1, value='DESCRIPTIVE STATISTICS').font = SECTION_TITLE_FONT
        ws.row_dimensions[current_row].height = 15.75
        current_row += 1

        desc_headers = ['Condition', 'N', 'Mean', 'Median', 'SD', 'SEM', 'Min', 'Max']
        for col_idx, header in enumerate(desc_headers, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
        current_row += 1

        for cond in conditions:
            values = data.get(cond, [])
            if values:
                arr = np.array(values)
                n = len(arr)
                mean = np.mean(arr)
                median = np.median(arr)
                std = np.std(arr, ddof=1)
                sem = std / np.sqrt(n) if n > 0 else 0

                row_values = [cond, n, round(mean, 3), round(median, 3),
                              round(std, 3), round(sem, 3), round(np.min(arr), 3), round(np.max(arr), 3)]
                for col_idx, val in enumerate(row_values, 1):
                    ws.cell(row=current_row, column=col_idx, value=val)
                current_row += 1

        current_row = 19

        # Section 3: PAIRWISE COMPARISONS
        ws.cell(row=current_row, column=1, value='PAIRWISE COMPARISONS').font = SECTION_TITLE_FONT
        ws.row_dimensions[current_row].height = 15.75
        current_row += 1

        pairwise_headers = ['Group 1', 'Group 2', 'Test', 'Statistic', 'P-value', 'Significant']
        for col_idx, header in enumerate(pairwise_headers, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
        current_row += 1

        # Get pairwise comparisons
        try:
            result = self.stats.auto_compare(stats_df, 'value', 'condition')
            post_hoc = result.get('post_hoc_tests', [])

            if post_hoc:
                for test in post_hoc:
                    test_name = 'Mann-Whitney U' if not result.get('is_parametric', True) else 'Tukey HSD'
                    significant = 'Yes' if test.p_value < 0.05 else 'No'

                    row_values = [test.group1, test.group2, test_name,
                                  round(test.statistic, 4) if hasattr(test, 'statistic') else '',
                                  f'{test.p_value:.4e}', significant]
                    for col_idx, val in enumerate(row_values, 1):
                        ws.cell(row=current_row, column=col_idx, value=val)
                    current_row += 1
        except Exception:
            pass
