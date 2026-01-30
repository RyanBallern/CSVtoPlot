"""Excel file importer for neuromorphology data (XLS and XLSX)."""

from pathlib import Path
from typing import List, Dict, Optional
import pandas as pd


class ExcelImporter:
    """Imports data from Excel files (XLS and XLSX)."""

    @staticmethod
    def import_file(
        file_path: Path,
        selected_parameters: Optional[List[str]] = None,
        sheet_index: int = 0
    ) -> pd.DataFrame:
        """
        Import data from an Excel file (XLS or XLSX).

        Args:
            file_path: Path to Excel file
            selected_parameters: List of parameters to import (None = all)
            sheet_index: Sheet index to read (default: 0 = first sheet)

        Returns:
            DataFrame with imported data

        Raises:
            FileNotFoundError: If file doesn't exist
            ValueError: If selected parameters not found in file
            ImportError: If required package (xlrd/openpyxl) not installed
        """
        file_path = Path(file_path)
        if not file_path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")

        ext = file_path.suffix.lower()

        # Read based on file extension
        if ext == '.xls':
            df = ExcelImporter._import_xls(file_path, sheet_index)
        elif ext == '.xlsx':
            df = ExcelImporter._import_xlsx(file_path, sheet_index)
        else:
            raise ValueError(f"Unsupported Excel format: {ext}")

        # Filter to selected parameters if specified
        if selected_parameters is not None:
            # Verify all selected parameters exist
            missing = set(selected_parameters) - set(df.columns)
            if missing:
                raise ValueError(f"Parameters not found in file: {missing}")

            # Select only the requested columns
            df = df[selected_parameters]

        return df

    @staticmethod
    def _import_xls(file_path: Path, sheet_index: int = 0) -> pd.DataFrame:
        """
        Import XLS file using xlrd.

        Args:
            file_path: Path to XLS file
            sheet_index: Sheet index to read

        Returns:
            DataFrame with all data from the sheet

        Raises:
            ImportError: If xlrd not installed
        """
        try:
            import xlrd
        except ImportError:
            raise ImportError(
                "xlrd package required for XLS files. "
                "Install with: pip install xlrd"
            )

        # Open workbook
        workbook = xlrd.open_workbook(file_path)
        sheet = workbook.sheet_by_index(sheet_index)

        # Extract headers (first row)
        headers = [sheet.cell_value(0, col) for col in range(sheet.ncols)]

        # Extract data (remaining rows)
        data = []
        for row_idx in range(1, sheet.nrows):
            row_data = {}
            for col_idx, header in enumerate(headers):
                if header:  # Skip empty headers
                    row_data[header] = sheet.cell_value(row_idx, col_idx)
            data.append(row_data)

        return pd.DataFrame(data)

    @staticmethod
    def _import_xlsx(file_path: Path, sheet_index: int = 0) -> pd.DataFrame:
        """
        Import XLSX file using openpyxl.

        Args:
            file_path: Path to XLSX file
            sheet_index: Sheet index to read

        Returns:
            DataFrame with all data from the sheet

        Raises:
            ImportError: If openpyxl not installed
        """
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise ImportError(
                "openpyxl package required for XLSX files. "
                "Install with: pip install openpyxl"
            )

        # Use pandas read_excel which handles openpyxl internally
        # This is more efficient than manually parsing
        df = pd.read_excel(file_path, sheet_name=sheet_index, engine='openpyxl')

        return df

    @staticmethod
    def import_file_as_dict(
        file_path: Path,
        selected_parameters: Optional[List[str]] = None,
        sheet_index: int = 0
    ) -> List[Dict[str, any]]:
        """
        Import data from Excel file as list of dictionaries.

        Args:
            file_path: Path to Excel file
            selected_parameters: List of parameters to import (None = all)
            sheet_index: Sheet index to read (default: 0)

        Returns:
            List of dictionaries, one per row

        Raises:
            FileNotFoundError: If file doesn't exist
            ValueError: If selected parameters not found in file
        """
        df = ExcelImporter.import_file(file_path, selected_parameters, sheet_index)
        return df.to_dict('records')

    @staticmethod
    def get_row_count(file_path: Path, sheet_index: int = 0) -> int:
        """
        Get number of data rows in Excel file (excluding header).

        Args:
            file_path: Path to Excel file
            sheet_index: Sheet index to read (default: 0)

        Returns:
            Number of data rows
        """
        df = ExcelImporter.import_file(file_path, sheet_index=sheet_index)
        return len(df)

    @staticmethod
    def get_sheet_names(file_path: Path) -> List[str]:
        """
        Get list of sheet names in Excel file.

        Args:
            file_path: Path to Excel file

        Returns:
            List of sheet names
        """
        file_path = Path(file_path)
        ext = file_path.suffix.lower()

        if ext == '.xls':
            import xlrd
            workbook = xlrd.open_workbook(file_path)
            return workbook.sheet_names()
        elif ext == '.xlsx':
            from openpyxl import load_workbook
            workbook = load_workbook(file_path, read_only=True)
            names = workbook.sheetnames
            workbook.close()
            return names
        else:
            raise ValueError(f"Unsupported Excel format: {ext}")
